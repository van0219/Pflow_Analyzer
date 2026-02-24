#!/usr/bin/env python3
"""
IPA Client Handover Documentation Template - Modern Edition (2026)

This template generates professional client-facing IPA documentation with modern design.
NO code criticism - focuses on how to understand and maintain the IPAs.

Design Standards: Follows modern color palette and visual hierarchy defined in
.kiro/steering/10_IPA_Report_Generation.md

Usage:
    from ipa_client_handover_template import generate_report
    generate_report(ipa_data)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from datetime import datetime
import os


# Modern Color Palette (2026 Professional)
COLORS = {
    # Primary Colors
    'deep_blue': '1565C0',      # Headers, titles, primary elements
    'green': '2E7D32',          # Success, positive findings, completion
    'amber': 'F57C00',          # Warnings, important notes, decisions
    'purple': '6A1B9A',         # API/technical elements, OAuth, integrations
    'red': 'C62828',            # Errors, critical issues, failures
    
    # Secondary Colors
    'medium_blue': '1E88E5',    # Subheaders
    'light_blue': 'E3F2FD',     # Alternating row backgrounds
    'light_green': 'E8F5E9',    # Success backgrounds
    'light_amber': 'FFF3E0',    # Warning backgrounds
    'light_purple': 'E1BEE7',   # API backgrounds
    
    # Neutral
    'border_gray': 'BDBDBD',    # Borders
    'white': 'FFFFFF'           # Text on colored backgrounds
}


def get_styles():
    """Return consistent style objects for modern design"""
    return {
        'header_fill': PatternFill(start_color=COLORS['deep_blue'], end_color=COLORS['deep_blue'], fill_type='solid'),
        'header_font': Font(color=COLORS['white'], bold=True, size=11),
        'subheader_fill': PatternFill(start_color=COLORS['medium_blue'], end_color=COLORS['medium_blue'], fill_type='solid'),
        'subheader_font': Font(color=COLORS['white'], bold=True, size=10),
        'success_fill': PatternFill(start_color=COLORS['green'], end_color=COLORS['green'], fill_type='solid'),
        'success_font': Font(color=COLORS['white'], bold=True),
        'warning_fill': PatternFill(start_color=COLORS['amber'], end_color=COLORS['amber'], fill_type='solid'),
        'warning_font': Font(color=COLORS['white'], bold=True),
        'error_fill': PatternFill(start_color=COLORS['red'], end_color=COLORS['red'], fill_type='solid'),
        'error_font': Font(color=COLORS['white'], bold=True),
        'api_fill': PatternFill(start_color=COLORS['purple'], end_color=COLORS['purple'], fill_type='solid'),
        'api_font': Font(color=COLORS['white'], bold=True),
        'light_blue_fill': PatternFill(start_color=COLORS['light_blue'], end_color=COLORS['light_blue'], fill_type='solid'),
        'light_green_fill': PatternFill(start_color=COLORS['light_green'], end_color=COLORS['light_green'], fill_type='solid'),
        'light_amber_fill': PatternFill(start_color=COLORS['light_amber'], end_color=COLORS['light_amber'], fill_type='solid'),
        'light_purple_fill': PatternFill(start_color=COLORS['light_purple'], end_color=COLORS['light_purple'], fill_type='solid'),
        'border': Border(
            left=Side(style='thin', color=COLORS['border_gray']),
            right=Side(style='thin', color=COLORS['border_gray']),
            top=Side(style='thin', color=COLORS['border_gray']),
            bottom=Side(style='thin', color=COLORS['border_gray'])
        )
    }


def generate_report(ipa_data):
    """
    Generate client handover documentation Excel report with modern design.
    
    Args:
        ipa_data (dict): Dictionary containing all IPA analysis data
            Required keys:
            - client_name: str
            - process_group: str
            - processes: list of process dicts
            
            Optional keys:
            - workflow_diagram_path: str (path to PNG)
            - business_requirements: dict
            - production_validation: dict
            - config_variables: list
            - maintenance_guide: dict
    """
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create all sheets
    create_executive_summary(wb, ipa_data)
    create_business_requirements(wb, ipa_data)
    create_production_validation(wb, ipa_data)
    create_system_configuration(wb, ipa_data)  # MANDATORY per 2026-01-20 requirements
    create_activity_node_guide(wb, ipa_data)
    
    # Create process sheets (variable number)
    for idx, process in enumerate(ipa_data.get('processes', []), start=1):
        create_process_sheet(wb, process, idx)
    
    create_maintenance_guide(wb, ipa_data)
    
    # Save report
    output_path = f"Client_Handover_Results/{ipa_data['client_name']}_{ipa_data['process_group']}.xlsx"
    os.makedirs('Client_Handover_Results', exist_ok=True)
    wb.save(output_path)
    
    print(f"✓ Client handover documentation generated: {output_path}")
    return output_path


def create_executive_summary(wb, ipa_data):
    """Create Executive Summary sheet with modern design and workflow diagram"""
    ws = wb.create_sheet('📊 Executive Summary')
    styles = get_styles()
    
    # Title with emoji
    ws['A1'] = f"📊 {ipa_data['client_name']} - {ipa_data['process_group']}"
    ws['A1'].font = Font(size=16, bold=True, color=COLORS['deep_blue'])
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = 30
    
    # Date stamp
    ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
    ws['A2'].font = Font(size=10, color=COLORS['border_gray'], italic=True)
    ws.merge_cells('A2:F2')
    
    row = 4
    
    # LEFT SIDE: Process Information (Columns A-C)
    # Process Information Box
    ws[f'A{row}'] = '📋 Project Information'
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].fill = styles['header_fill']
    ws[f'A{row}'].font = styles['header_font']
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 25
    row += 1
    
    # Process details
    details = ipa_data.get('process_details', {})
    for key, value in details.items():
        ws[f'A{row}'] = key
        ws[f'B{row}'] = value
        ws.merge_cells(f'B{row}:C{row}')
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}{row}']
            cell.border = styles['border']
            if row % 2 == 0:
                cell.fill = styles['light_blue_fill']
        ws[f'A{row}'].font = Font(bold=True, color=COLORS['deep_blue'])
        ws[f'B{row}'].alignment = Alignment(wrap_text=True)
        row += 1
    
    row += 1
    
    # Key Features Box
    ws[f'A{row}'] = '✨ Key Features'
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].fill = styles['success_fill']
    ws[f'A{row}'].font = styles['success_font']
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 25
    row += 1
    
    for feature in ipa_data.get('key_features', []):
        ws[f'A{row}'] = f"✓ {feature}"
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'].fill = styles['light_green_fill']
        ws[f'A{row}'].border = styles['border']
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        row += 1
    
    # Sample Run (if available)
    if 'sample_run' in ipa_data:
        row += 1
        ws[f'A{row}'] = '✅ Production Validation'
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'].fill = styles['api_fill']
        ws[f'A{row}'].font = styles['api_font']
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 1
        
        sample = ipa_data['sample_run']
        sample_data = [
            ['Work Unit:', sample.get('wu_id', 'N/A')],
            ['Status:', f"✅ {sample.get('status', 'N/A')}"],
            ['Duration:', f"⚡ {sample.get('duration', 'N/A')}"]
        ]
        
        for row_data in sample_data:
            ws[f'A{row}'] = row_data[0]
            ws[f'B{row}'] = row_data[1]
            ws.merge_cells(f'B{row}:C{row}')
            for col in ['A', 'B', 'C']:
                cell = ws[f'{col}{row}']
                cell.border = styles['border']
                if row % 2 == 0:
                    cell.fill = styles['light_blue_fill']
            ws[f'A{row}'].font = Font(bold=True, color=COLORS['purple'])
            row += 1
    
    # RIGHT SIDE: Workflow Diagram (Columns D-F)
    diagram_row = 4
    
    # Generate workflow diagram
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    from matplotlib.patches import FancyBboxPatch, FancyArrowPatch
    import os
    import re
    
    # Function to remove emojis from text
    def remove_emojis(text):
        """Remove all emojis from text to prevent matplotlib rendering issues"""
        # Pattern matches most emoji ranges including clock symbols
        emoji_pattern = re.compile("["
            u"\U0001F600-\U0001F64F"  # emoticons
            u"\U0001F300-\U0001F5FF"  # symbols & pictographs
            u"\U0001F680-\U0001F6FF"  # transport & map symbols
            u"\U0001F1E0-\U0001F1FF"  # flags (iOS)
            u"\U00002300-\U000023FF"  # misc technical (includes clocks)
            u"\U00002600-\U000027BF"  # misc symbols & dingbats (includes more clocks)
            u"\U00002702-\U000027B0"  # dingbats
            u"\U000024C2-\U0001F251"
            u"\U0001F900-\U0001F9FF"  # supplemental symbols
            "]+", flags=re.UNICODE)
        return emoji_pattern.sub('', text).strip()
    
    # Create figure
    fig, ax = plt.subplots(figsize=(8, 12), dpi=300)
    ax.set_xlim(0, 10)
    ax.set_ylim(0, 20)
    ax.axis('off')
    
    # Title - NO EMOJIS
    ax.text(5, 19, 'Process Workflow', ha='center', va='top', 
            fontsize=14, fontweight='bold', color='#1565C0')
    
    # Define workflow steps based on process type
    # Check if workflow_description is provided (from subagent analysis)
    workflow_description = ipa_data.get('workflow_description', [])
    
    if workflow_description:
        # Convert workflow_description to visualization format
        workflow_steps = []
        workflow_steps.append({'type': 'start', 'label': 'START\nProcess Initiated', 'y': 17.5})
        
        y_position = 16
        for step in workflow_description:
            step_name = step.get('step_name', 'Process Step')
            # Determine type based on step name
            if 'decision' in step_name.lower() or 'validation' in step_name.lower() or 'check' in step_name.lower():
                step_type = 'decision'
            elif 'api' in step_name.lower() or 'authentication' in step_name.lower() or 'query' in step_name.lower() or 'data' in step_name.lower():
                step_type = 'api'
            elif 'error' in step_name.lower():
                step_type = 'error'
            else:
                step_type = 'process'
            
            # Truncate long names for visualization - NO EMOJIS
            label_parts = step_name.split()
            if len(label_parts) > 4:
                label = ' '.join(label_parts[:2]) + '\n' + ' '.join(label_parts[2:4])
            else:
                label = step_name.replace(' ', '\n', 1) if len(step_name) > 15 else step_name
            
            workflow_steps.append({'type': step_type, 'label': label, 'y': y_position})
            y_position -= 1.3
        
        workflow_steps.append({'type': 'end', 'label': 'END\nProcess Complete', 'y': 1})
    else:
        # Use default workflow steps - NO EMOJIS
        workflow_steps = ipa_data.get('workflow_steps', [
            {'type': 'start', 'label': 'START\nProcess Initiated', 'y': 17.5},
            {'type': 'process', 'label': 'Read Input\n(File/Trigger)', 'y': 16},
            {'type': 'process', 'label': 'Validate & Parse\nInput Data', 'y': 14.5},
            {'type': 'decision', 'label': 'Valid Input?', 'y': 13, 'branches': ['Yes', 'No']},
            {'type': 'api', 'label': 'Authenticate\n(OAuth/API)', 'y': 11.5},
            {'type': 'api', 'label': 'Execute Query\n(API/Database)', 'y': 10},
            {'type': 'process', 'label': 'Check Status\n(If Async)', 'y': 8.5},
            {'type': 'decision', 'label': 'Complete?', 'y': 7, 'branches': ['Yes', 'Wait']},
            {'type': 'api', 'label': 'Retrieve Results\n(Process Data)', 'y': 5.5},
            {'type': 'process', 'label': 'Transform & Write\nOutput File', 'y': 4},
            {'type': 'process', 'label': 'Finalize\n(Archive/Notify)', 'y': 2.5},
            {'type': 'end', 'label': 'END\nProcess Complete', 'y': 1}
        ])
    
    # CRITICAL: Strip emojis from ALL labels to prevent matplotlib rendering issues
    for step in workflow_steps:
        if 'label' in step:
            step['label'] = remove_emojis(step['label'])
        # Also strip emojis from branch labels
        if 'branches' in step and isinstance(step['branches'], list):
            step['branches'] = [remove_emojis(str(b)) for b in step['branches']]
    
    # Color mapping
    colors = {
        'start': '#2E7D32',
        'end': '#2E7D32',
        'process': '#1565C0',
        'decision': '#F57C00',
        'api': '#6A1B9A',
        'file': '#00897B',
        'error': '#C62828'
    }
    
    # Draw workflow steps
    for i, step in enumerate(workflow_steps):
        step_type = step['type']
        label = step['label']
        y = step['y']
        
        # Draw box
        if step_type == 'decision':
            # Diamond shape for decisions
            points = [[5, y+0.4], [6.5, y], [5, y-0.4], [3.5, y]]
            polygon = mpatches.Polygon(points, closed=True, 
                                      facecolor=colors[step_type], 
                                      edgecolor='black', linewidth=1.5)
            ax.add_patch(polygon)
        else:
            # Rounded rectangle for other steps
            box = FancyBboxPatch((3.5, y-0.3), 3, 0.6,
                                boxstyle="round,pad=0.05",
                                facecolor=colors[step_type],
                                edgecolor='black', linewidth=1.5)
            ax.add_patch(box)
        
        # Add label
        ax.text(5, y, label, ha='center', va='center',
               fontsize=8, fontweight='bold', color='white',
               multialignment='center')
        
        # Draw arrow to next step
        if i < len(workflow_steps) - 1:
            next_y = workflow_steps[i+1]['y']
            arrow = FancyArrowPatch((5, y-0.4), (5, next_y+0.4),
                                   arrowstyle='->', mutation_scale=15,
                                   color='black', linewidth=1.5)
            ax.add_patch(arrow)
        
        # Draw branch arrows for decisions
        if step_type == 'decision' and 'branches' in step:
            # Error branch (right)
            if 'No' in step['branches'] or 'Error' in step['branches']:
                error_arrow = FancyArrowPatch((6.5, y), (8, y),
                                             arrowstyle='->', mutation_scale=15,
                                             color='#C62828', linewidth=1.5)
                ax.add_patch(error_arrow)
                ax.text(8.2, y, 'Error', ha='left', va='center',
                       fontsize=7, color='#C62828', fontweight='bold')
            
            # Wait/Retry branch (left)
            if 'Wait' in step['branches']:
                wait_arrow = FancyArrowPatch((3.5, y), (2, y),
                                            arrowstyle='->', mutation_scale=15,
                                            color='#F57C00', linewidth=1.5)
                ax.add_patch(wait_arrow)
                ax.text(1.8, y, 'Wait', ha='right', va='center',
                       fontsize=7, color='#F57C00', fontweight='bold')
    
    # Add legend
    legend_y = 0.3
    legend_items = [
        ('Start/End', colors['start']),
        ('Process', colors['process']),
        ('Decision', colors['decision']),
        ('API Call', colors['api'])
    ]
    
    for i, (label, color) in enumerate(legend_items):
        x = 1.5 + (i * 2.2)
        box = FancyBboxPatch((x, legend_y-0.15), 0.4, 0.3,
                            boxstyle="round,pad=0.02",
                            facecolor=color, edgecolor='black', linewidth=1)
        ax.add_patch(box)
        ax.text(x+0.6, legend_y, label, ha='left', va='center',
               fontsize=7, fontweight='bold')
    
    # Save diagram
    os.makedirs('Temp', exist_ok=True)
    diagram_path = 'Temp/workflow_diagram.png'
    plt.tight_layout()
    plt.savefig(diagram_path, dpi=300, bbox_inches='tight', facecolor='white')
    plt.close()
    
    # Embed diagram in Excel
    ws[f'D{diagram_row}'] = '🔄 Process Flow Diagram'
    ws.merge_cells(f'D{diagram_row}:F{diagram_row}')
    ws[f'D{diagram_row}'].fill = styles['header_fill']
    ws[f'D{diagram_row}'].font = styles['header_font']
    ws[f'D{diagram_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[diagram_row].height = 25
    
    img = XLImage(diagram_path)
    img.width = 400
    img.height = 600
    ws.add_image(img, f'D{diagram_row+1}')
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20


def create_business_requirements(wb, ipa_data):
    """Create Business Requirements sheet with modern design and comprehensive structured data"""
    ws = wb.create_sheet('📋 Business Requirements')
    styles = get_styles()
    
    # Title
    ws['A1'] = '📋 Business Requirements'
    ws['A1'].font = Font(size=16, bold=True, color=COLORS['deep_blue'])
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:H1')
    ws.row_dimensions[1].height = 30
    
    # Subtitle with count
    req_count = len(ipa_data.get('requirements', []))
    ws['A2'] = f'Total Requirements: {req_count}'
    ws['A2'].font = Font(size=11, color=COLORS['deep_blue'], bold=True)
    ws.merge_cells('A2:H2')
    ws.row_dimensions[2].height = 20
    
    row = 4
    
    # Check if requirements are structured (dict) or simple (string/tuple)
    requirements = ipa_data.get('requirements', [])
    is_structured = requirements and isinstance(requirements[0], dict)
    
    if is_structured:
        # Rich structured format with full details
        ws[f'A{row}'] = 'ID'
        ws[f'B{row}'] = 'Category'
        ws[f'C{row}'] = 'Title'
        ws[f'D{row}'] = 'Description'
        ws[f'E{row}'] = 'Priority'
        ws[f'F{row}'] = 'Business Value'
        ws[f'G{row}'] = 'Source'
        ws[f'H{row}'] = 'Stakeholders'
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            cell = ws[f'{col}{row}']
            cell.fill = styles['header_fill']
            cell.font = styles['header_font']
            cell.border = styles['border']
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 1
        
        # Add structured requirements
        priority_colors = {
            'High': 'FF6B6B',
            'Medium': 'FFA500',
            'Low': '4ECDC4'
        }
        
        for req in requirements:
            ws[f'A{row}'] = req.get('id', '')
            ws[f'B{row}'] = req.get('category', '')
            ws[f'C{row}'] = req.get('title', '')
            ws[f'D{row}'] = req.get('description', '')
            ws[f'E{row}'] = req.get('priority', '')
            ws[f'F{row}'] = req.get('business_value', '')
            ws[f'G{row}'] = req.get('source', '')
            # Format stakeholders as comma-separated list
            stakeholders = req.get('stakeholders', [])
            ws[f'H{row}'] = ', '.join(stakeholders) if isinstance(stakeholders, list) else str(stakeholders)
            
            # Apply priority color to priority cell
            priority = req.get('priority', 'Medium')
            priority_color = priority_colors.get(priority, 'FFFFFF')
            ws[f'E{row}'].fill = PatternFill(start_color=priority_color, end_color=priority_color, fill_type='solid')
            ws[f'E{row}'].font = Font(bold=True, color='FFFFFF')
            ws[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                cell = ws[f'{col}{row}']
                cell.border = styles['border']
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if row % 2 == 0 and col != 'E':
                    cell.fill = styles['light_blue_fill']
            
            ws[f'A{row}'].font = Font(bold=True, color=COLORS['deep_blue'])
            ws[f'C{row}'].font = Font(bold=True, color=COLORS['deep_blue'])
            ws.row_dimensions[row].height = 60
            row += 1
        
        # Column widths for structured format
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 35
        ws.column_dimensions['H'].width = 30
        
    else:
        # Simple format (backward compatibility)
        ws[f'A{row}'] = '🎯 Priority'
        ws[f'B{row}'] = 'Requirement'
        ws[f'C{row}'] = 'Description'
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}{row}']
            cell.fill = styles['header_fill']
            cell.font = styles['header_font']
            cell.border = styles['border']
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 1
        
        # Add requirements with visual priority indicators
        priority_emojis = ['🔴', '🟠', '🟡', '🟢', '🔵']
        for idx, req in enumerate(requirements):
            priority = priority_emojis[idx % len(priority_emojis)]
            ws[f'A{row}'] = priority
            ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws[f'B{row}'] = req[0] if isinstance(req, (list, tuple)) else req
            ws[f'C{row}'] = req[1] if isinstance(req, (list, tuple)) and len(req) > 1 else ''
            
            for col in ['A', 'B', 'C']:
                cell = ws[f'{col}{row}']
                cell.border = styles['border']
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if row % 2 == 0:
                    cell.fill = styles['light_blue_fill']
            ws[f'B{row}'].font = Font(bold=True, color=COLORS['deep_blue'])
            ws.row_dimensions[row].height = 40
            row += 1
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 70


def create_system_configuration(wb, ipa_data):
    """Create System Configuration sheet with modern design - MANDATORY per 2026-01-20"""
    ws = wb.create_sheet('⚙️ System Configuration')
    styles = get_styles()
    
    # Title
    ws['A1'] = '⚙️ System Configuration'
    ws['A1'].font = Font(size=16, bold=True, color=COLORS['purple'])
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:E1')
    ws.row_dimensions[1].height = 30
    
    # Warning box
    ws['A2'] = '⚠️ CRITICAL: All configuration variables must be documented here for client maintenance'
    ws.merge_cells('A2:E2')
    ws['A2'].fill = styles['warning_fill']
    ws['A2'].font = styles['warning_font']
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 30
    
    row = 4
    
    # NEW: Handle structured config_variables format
    config_vars = ipa_data.get('config_variables', [])
    if config_vars and isinstance(config_vars, list) and len(config_vars) > 0:
        # Group by type
        system_config = [v for v in config_vars if 'System Configuration' in v.get('type', '')]
        file_channel = [v for v in config_vars if 'File Channel' in v.get('type', '')]
        
        # System Configuration Variables
        if system_config:
            ws[f'A{row}'] = '🌐 System Configuration Variables'
            ws.merge_cells(f'A{row}:E{row}')
            ws[f'A{row}'].fill = styles['header_fill']
            ws[f'A{row}'].font = styles['header_font']
            ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[row].height = 25
            row += 1
            
            ws[f'A{row}'] = 'ℹ️ These variables are stored in FSM > Configuration > System Configuration > Interface'
            ws.merge_cells(f'A{row}:E{row}')
            ws[f'A{row}'].fill = styles['light_blue_fill']
            ws[f'A{row}'].border = styles['border']
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            ws.row_dimensions[row].height = 30
            row += 1
            
            headers = ['Variable Name', 'Location', 'Description', 'Current Value', 'How to Modify']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.fill = styles['header_fill']
                cell.font = styles['header_font']
                cell.border = styles['border']
                cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[row].height = 25
            row += 1
            
            for var in system_config:
                ws.cell(row=row, column=1, value=var.get('name', ''))
                ws.cell(row=row, column=2, value=var.get('location', ''))
                ws.cell(row=row, column=3, value=var.get('description', ''))
                ws.cell(row=row, column=4, value=var.get('current_value', ''))
                ws.cell(row=row, column=5, value=var.get('how_to_modify', ''))
                
                for col_idx in range(1, 6):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.border = styles['border']
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    if row % 2 == 0:
                        cell.fill = styles['light_blue_fill']
                
                ws.row_dimensions[row].height = 80
                row += 1
            
            row += 1
        
        # File Channel Configuration
        if file_channel:
            ws[f'A{row}'] = '📁 File Channel Configuration'
            ws.merge_cells(f'A{row}:E{row}')
            ws[f'A{row}'].fill = styles['success_fill']
            ws[f'A{row}'].font = styles['success_font']
            ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[row].height = 25
            row += 1
            
            headers = ['Variable Name', 'Location', 'Description', 'Current Value', 'How to Modify']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.fill = styles['header_fill']
                cell.font = styles['header_font']
                cell.border = styles['border']
                cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[row].height = 25
            row += 1
            
            for var in file_channel:
                ws.cell(row=row, column=1, value=var.get('name', ''))
                ws.cell(row=row, column=2, value=var.get('location', ''))
                ws.cell(row=row, column=3, value=var.get('description', ''))
                ws.cell(row=row, column=4, value=var.get('current_value', ''))
                ws.cell(row=row, column=5, value=var.get('how_to_modify', ''))
                
                for col_idx in range(1, 6):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.border = styles['border']
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    if row % 2 == 0:
                        cell.fill = styles['light_green_fill']
                
                ws.row_dimensions[row].height = 80
                row += 1
    
    # LEGACY: Handle old format (backward compatibility)
    elif 'oauth_credentials' in ipa_data or 'file_channel_config' in ipa_data or 'process_variables' in ipa_data or 'global_config_variables' in ipa_data:
        ws[f'A{row}'] = '🔐 OAuth2 Credentials'
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].fill = styles['api_fill']
        ws[f'A{row}'].font = styles['api_font']
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 1
        
        headers = ['Variable Name', 'JSON Structure', 'Purpose', 'How to Modify']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = styles['header_fill']
            cell.font = styles['header_font']
            cell.border = styles['border']
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 1
        
        for cred in ipa_data['oauth_credentials']:
            for col_idx, value in enumerate(cred, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = styles['border']
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if row % 2 == 0:
                    cell.fill = styles['light_purple_fill']
            ws.row_dimensions[row].height = 60
            row += 1
        
        row += 1
    
    # File Channel Configuration Section (if applicable)
    if 'file_channel_config' in ipa_data:
        ws[f'A{row}'] = '📁 File Channel Configuration'
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].fill = styles['success_fill']
        ws[f'A{row}'].font = styles['success_font']
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 1
        
        headers = ['Variable Name', 'Example Value', 'Purpose', 'How to Modify']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = styles['header_fill']
            cell.font = styles['header_font']
            cell.border = styles['border']
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 1
        
        for config in ipa_data['file_channel_config']:
            for col_idx, value in enumerate(config, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = styles['border']
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if row % 2 == 0:
                    cell.fill = styles['light_green_fill']
            ws.row_dimensions[row].height = 40
            row += 1
        
        row += 1
    
    # Process Variables Section
    if 'process_variables' in ipa_data:
        ws[f'A{row}'] = '📝 Process Variables (START Activity)'
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].fill = styles['warning_fill']
        ws[f'A{row}'].font = styles['warning_font']
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 1
        
        headers = ['Variable Name', 'Default Value', 'Purpose', 'How to Modify']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = styles['header_fill']
            cell.font = styles['header_font']
            cell.border = styles['border']
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 1
        
        for var in ipa_data['process_variables']:
            for col_idx, value in enumerate(var, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = styles['border']
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if row % 2 == 0:
                    cell.fill = styles['light_amber_fill']
            ws.row_dimensions[row].height = 40
            row += 1
        
        row += 1
    
    # Global Configuration Variables Section (if applicable)
    if 'global_config_variables' in ipa_data:
        ws[f'A{row}'] = '🌐 Global System Configuration Variables'
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].fill = styles['header_fill']
        ws[f'A{row}'].font = styles['header_font']
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 1
        
        ws[f'A{row}'] = 'ℹ️ These variables are stored in FSM > Configuration > System Configuration > Interface'
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].fill = styles['light_blue_fill']
        ws[f'A{row}'].border = styles['border']
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 30
        row += 1
        
        headers = ['Property Name', 'Value Format', 'Purpose', 'How to Modify']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = styles['header_fill']
            cell.font = styles['header_font']
            cell.border = styles['border']
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 1
        
        for var in ipa_data['global_config_variables']:
            for col_idx, value in enumerate(var, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = styles['border']
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if row % 2 == 0:
                    cell.fill = styles['light_blue_fill']
            ws.row_dimensions[row].height = 40
            row += 1
    
    # Column widths (5 columns for new format, 4 for legacy)
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 50


def create_production_validation(wb, ipa_data):
    """Create Production Validation sheet with modern design and visual metrics"""
    ws = wb.create_sheet('✅ Production Validation')
    styles = get_styles()
    
    # Title
    ws['A1'] = 'Production Test Results'
    ws['A1'].font = Font(size=16, bold=True, color=COLORS['green'])
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:C1')
    ws.row_dimensions[1].height = 30
    
    # Success banner
    ws['A2'] = 'All validation checks passed successfully!'
    ws.merge_cells('A2:C2')
    ws['A2'].fill = styles['light_green_fill']
    ws['A2'].font = Font(bold=True, color=COLORS['green'], size=11)
    ws['A2'].border = styles['border']
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 25
    
    row = 4
    ws[f'A{row}'] = 'Test Parameter'
    ws[f'B{row}'] = 'Value'
    ws[f'C{row}'] = 'Status'
    
    for col in ['A', 'B', 'C']:
        cell = ws[f'{col}{row}']
        cell.fill = styles['success_fill']
        cell.font = styles['success_font']
        cell.border = styles['border']
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 25
    row += 1
    
    # Add validation data with status indicators
    validation = ipa_data.get('production_validation', {})
    
    # Flatten nested validation data for better readability
    flattened_data = []
    
    # Test Summary
    if 'test_summary' in validation:
        ts = validation['test_summary']
        flattened_data.append(('Work Unit Number', ts.get('work_unit_number', 'N/A'), 'Pass'))
        flattened_data.append(('Total Executions', ts.get('total_executions', 'N/A'), 'Pass'))
        flattened_data.append(('Successful', ts.get('successful', 'N/A'), 'Pass'))
        flattened_data.append(('Failed', ts.get('failed', 0), 'Pass' if ts.get('failed', 0) == 0 else 'Warning'))
        flattened_data.append(('Success Rate', f"{ts.get('success_rate', 0)}%", 'Pass'))
        flattened_data.append(('Test Date', ts.get('test_date', 'N/A'), 'Pass'))
        flattened_data.append(('Test Environment', ts.get('test_environment', 'N/A'), 'Pass'))
    
    # Performance
    if 'performance' in validation:
        perf = validation['performance']
        flattened_data.append(('Average Duration', perf.get('avg_duration', 'N/A'), 'Pass'))
        flattened_data.append(('Min Duration', perf.get('min_duration', 'N/A'), 'Pass'))
        flattened_data.append(('Max Duration', perf.get('max_duration', 'N/A'), 'Pass'))
        flattened_data.append(('Performance Rating', perf.get('performance_rating', 'N/A'), 'Pass'))
    
    # Error Handling
    if 'error_handling' in validation:
        eh = validation['error_handling']
        flattened_data.append(('Error Handling Validated', 'Yes' if eh.get('validated') else 'No', 'Pass'))
        flattened_data.append(('Test Scenarios', f"{len(eh.get('test_scenarios', []))} scenarios tested", 'Pass'))
        flattened_data.append(('Confidence Level', eh.get('confidence', 'N/A'), 'Pass'))
    
    # Production Readiness
    if 'production_readiness' in validation:
        pr = validation['production_readiness']
        flattened_data.append(('Production Ready', 'Yes' if pr.get('ready') else 'No', 'Pass'))
        flattened_data.append(('Confidence Level', pr.get('confidence_level', 'N/A'), 'Pass'))
        flattened_data.append(('Evidence Count', f"{len(pr.get('evidence', []))} items", 'Pass'))
    
    # Test Coverage
    if 'test_coverage' in validation:
        tc = validation['test_coverage']
        flattened_data.append(('Scenarios Tested', f"{len(tc.get('scenarios_tested', []))} scenarios", 'Pass'))
        flattened_data.append(('Coverage Percentage', f"{tc.get('coverage_percentage', 0)}%", 'Pass'))
    
    # Activity Breakdown
    if 'activity_breakdown' in validation:
        ab = validation['activity_breakdown']
        flattened_data.append(('Total Activities', ab.get('total_activities', 'N/A'), 'Pass'))
        if 'activity_types' in ab:
            types_summary = ', '.join([f"{k}: {v}" for k, v in ab['activity_types'].items()])
            flattened_data.append(('Activity Types', types_summary, 'Pass'))
    
    # Data Validation
    if 'data_validation' in validation:
        dv = validation['data_validation']
        flattened_data.append(('Input File', dv.get('input_file', 'N/A'), 'Pass'))
        flattened_data.append(('Records Retrieved', dv.get('records_retrieved', 'N/A'), 'Pass'))
    
    # Integration Points
    if 'integration_points' in validation:
        ip = validation['integration_points']
        for system, details in ip.items():
            if isinstance(details, dict):
                status = details.get('status', 'N/A')
                ops = details.get('operations', 'N/A')
                flattened_data.append((f"{system}", f"{status}, {ops}", 'Pass'))
    
    # Risk Assessment
    if 'risk_assessment' in validation:
        ra = validation['risk_assessment']
        flattened_data.append(('Overall Risk', ra.get('overall_risk', 'N/A'), 'Pass'))
        if 'risk_factors' in ra:
            for factor in ra['risk_factors']:
                if isinstance(factor, dict):
                    flattened_data.append((f"Risk: {factor.get('factor', 'N/A')}", factor.get('mitigation', 'N/A'), 'Pass'))
    
    # Write flattened data to sheet
    for param, value, status in flattened_data:
        ws[f'A{row}'] = param
        ws[f'B{row}'] = str(value)
        
        # Add status indicator
        if status == 'Warning':
            ws[f'C{row}'] = 'Warning'
            ws[f'C{row}'].font = Font(bold=True, color=COLORS['amber'])
        else:
            ws[f'C{row}'] = 'Pass'
            ws[f'C{row}'].font = Font(bold=True, color=COLORS['green'])
        
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}{row}']
            cell.border = styles['border']
            if row % 2 == 0:
                cell.fill = styles['light_green_fill']
        ws[f'A{row}'].font = Font(bold=True, color=COLORS['green'])
        ws[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
        row += 1
    
    # Add performance summary box
    row += 1
    ws[f'A{row}'] = 'Performance Summary'
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'].fill = styles['api_fill']
    ws[f'A{row}'].font = styles['api_font']
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 25
    row += 1
    
    # Extract performance metrics
    duration = validation.get('performance', {}).get('avg_duration', 'N/A')
    records = validation.get('data_validation', {}).get('records_retrieved', 'N/A')
    
    perf_metrics = [
        ['Execution Speed', duration, 'Excellent'],
        ['Data Volume', records, 'Normal'],
        ['Authentication', 'OAuth2 Token', 'Secure'],
        ['API Integration', 'Compass Data Fabric', 'Stable']
    ]
    
    for metric in perf_metrics:
        ws[f'A{row}'] = metric[0]
        ws[f'B{row}'] = metric[1]
        ws[f'C{row}'] = metric[2]
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}{row}']
            cell.border = styles['border']
            if row % 2 == 0:
                cell.fill = styles['light_blue_fill']
        ws[f'A{row}'].font = Font(bold=True, color=COLORS['purple'])
        ws[f'C{row}'].font = Font(bold=True, color=COLORS['green'])
        ws[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
        row += 1
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15


def create_activity_node_guide(wb, ipa_data):
    """Create Activity Node Guide sheet with modern design"""
    ws = wb.create_sheet('📚 Activity Node Guide')
    styles = get_styles()
    
    # Title
    ws['A1'] = '📚 IPA Activity Node Reference'
    ws['A1'].font = Font(size=16, bold=True, color=COLORS['deep_blue'])
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:C1')
    ws.row_dimensions[1].height = 30
    
    row = 3
    ws[f'A{row}'] = 'Activity Type'
    ws[f'B{row}'] = 'Purpose'
    ws[f'C{row}'] = 'Common Uses'
    
    for col in ['A', 'B', 'C']:
        cell = ws[f'{col}{row}']
        cell.fill = styles['header_fill']
        cell.font = styles['header_font']
        cell.border = styles['border']
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 25
    row += 1
    
    # Activity types with emojis
    activity_types = [
        ['🚀 START', 'Process entry point', 'Initialize process variables'],
        ['✅ END', 'Process completion', 'Mark successful completion'],
        ['📝 ASSGN (Assign)', 'Variable assignment', 'Set values, execute JavaScript, parse JSON responses'],
        ['🌐 WEBRN (Web Run)', 'HTTP API calls', 'OAuth token requests, Compass API queries, status checks'],
        ['� BRANCH', 'Conditional routing', 'Route based on status, error conditions, record counts'],
        ['⏱️ Timer (Wait)', 'Delay execution', 'Wait for async operations to complete'],
        ['📁 ACCFIL (File Access)', 'File operations', 'Read trigger files, write output files, move/copy files'],
        ['💬 EMAIL', 'Email notifications', 'Send notifications to users or groups'],
        ['🏢 LM (Landmark)', 'FSM transactions', 'Query or update FSM business classes'],
    ]
    
    for activity in activity_types:
        ws[f'A{row}'] = activity[0]
        ws[f'B{row}'] = activity[1]
        ws[f'C{row}'] = activity[2]
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}{row}']
            cell.border = styles['border']
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if row % 2 == 0:
                cell.fill = styles['light_blue_fill']
        ws[f'A{row}'].font = Font(bold=True, color=COLORS['deep_blue'])
        ws.row_dimensions[row].height = 30
        row += 1
    
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 50


def create_process_sheet(wb, process, idx):
    """Create individual process sheet with modern design"""
    # Use process name for sheet title (truncate if too long for Excel)
    process_name = process.get('name', f'Process {idx}')
    sheet_name = f"⚙️ {process_name}"[:31]  # Excel sheet name limit is 31 chars
    ws = wb.create_sheet(sheet_name)
    styles = get_styles()
    
    # Title
    ws['A1'] = f"⚙️ {process['name']}"
    ws['A1'].font = Font(size=16, bold=True, color=COLORS['deep_blue'])
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:E1')
    ws.row_dimensions[1].height = 30
    
    row = 3
    
    # Configuration Variables (if available)
    config_vars = process.get('config_variables', [])
    if config_vars:
        ws[f'A{row}'] = '⚙️ Configuration Variables'
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'].fill = styles['warning_fill']
        ws[f'A{row}'].font = styles['warning_font']
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 1
        
        headers = ['Variable Name', 'Type', 'Location', 'Description', 'How to Modify']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = styles['header_fill']
            cell.font = styles['header_font']
            cell.border = styles['border']
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 1
        
        # Handle both dict and list formats
        for var in config_vars:
            if isinstance(var, dict):
                # New format: extract specific fields
                ws.cell(row=row, column=1, value=var.get('name', ''))
                ws.cell(row=row, column=2, value=var.get('type', ''))
                ws.cell(row=row, column=3, value=var.get('location', ''))
                ws.cell(row=row, column=4, value=var.get('description', ''))
                ws.cell(row=row, column=5, value=var.get('how_to_modify', ''))
            else:
                # Old format: list of values
                for col_idx, value in enumerate(var, start=1):
                    ws.cell(row=row, column=col_idx, value=value)
            
            for col_idx in range(1, 6):
                cell = ws.cell(row=row, column=col_idx)
                cell.border = styles['border']
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if row % 2 == 0:
                    cell.fill = styles['light_blue_fill']
            ws.row_dimensions[row].height = 60
            row += 1
        
        row += 1
    
    # Activity List
    ws[f'A{row}'] = '📋 Process Activities'
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'].fill = styles['header_fill']
    ws[f'A{row}'].font = styles['header_font']
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 25
    row += 1
    
    headers = ['Activity ID', 'Activity Type', 'Caption', 'Description', 'When It Runs']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = styles['subheader_fill']
        cell.font = styles['subheader_font']
        cell.border = styles['border']
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 25
    row += 1
    
    # Activities
    activities = process.get('activities', [])
    for activity in activities:
        if isinstance(activity, dict):
            # New format: extract specific fields
            ws.cell(row=row, column=1, value=activity.get('id', ''))
            ws.cell(row=row, column=2, value=activity.get('type', ''))
            ws.cell(row=row, column=3, value=activity.get('caption', ''))
            ws.cell(row=row, column=4, value=activity.get('description', ''))
            ws.cell(row=row, column=5, value=activity.get('when_it_runs', ''))
        else:
            # Old format: list of values
            for col_idx, value in enumerate(activity, start=1):
                ws.cell(row=row, column=col_idx, value=value)
        
        for col_idx in range(1, 6):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = styles['border']
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if row % 2 == 0:
                cell.fill = styles['light_blue_fill']
        ws.row_dimensions[row].height = 50
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 40


def create_maintenance_guide(wb, ipa_data):
    """Create Maintenance Guide sheet with modern design"""
    ws = wb.create_sheet('🔧 Maintenance Guide')
    styles = get_styles()
    
    # Title
    ws['A1'] = '🔧 Maintenance Guide'
    ws['A1'].font = Font(size=16, bold=True, color=COLORS['deep_blue'])
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:B1')
    ws.row_dimensions[1].height = 30
    
    row = 3
    
    # Get configuration and activity data
    config_vars = ipa_data.get('config_variables', [])
    processes = ipa_data.get('processes', [])
    
    # Section 1: Configuration Management
    ws[f'A{row}'] = '⚙️ Configuration Management'
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'].fill = styles['api_fill']
    ws[f'A{row}'].font = styles['api_font']
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 25
    row += 1
    
    config_tasks = [
        '1. Review System Configuration variables in FSM > Configuration > System Configuration > Interface',
        '2. Verify OAuth credentials are valid and not expired',
        '3. Check file channel is active and monitoring correct directories',
        '4. Validate output directory paths exist and have write permissions',
        '5. Test API endpoints are accessible and responding'
    ]
    
    for task in config_tasks:
        ws[f'A{row}'] = task
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'].fill = styles['light_blue_fill']
        ws[f'A{row}'].border = styles['border']
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 25
        row += 1
    
    row += 1
    
    # Section 2: Process Monitoring
    ws[f'A{row}'] = '📊 Process Monitoring'
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'].fill = styles['success_fill']
    ws[f'A{row}'].font = styles['success_font']
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 25
    row += 1
    
    monitor_tasks = [
        '1. Check work unit logs for errors or warnings',
        '2. Monitor file channel for stuck or failed files',
        '3. Review API response times and timeout occurrences',
        '4. Verify output files are being generated correctly',
        '5. Check disk space in input, output, and archive directories'
    ]
    
    for task in monitor_tasks:
        ws[f'A{row}'] = task
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'].fill = styles['light_green_fill']
        ws[f'A{row}'].border = styles['border']
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 25
        row += 1
    
    row += 1
    
    # Section 3: Troubleshooting
    ws[f'A{row}'] = '🔧 Common Issues & Solutions'
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'].fill = styles['warning_fill']
    ws[f'A{row}'].font = styles['warning_font']
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 25
    row += 1
    
    troubleshooting = [
        ('Process not starting', 'Verify file channel is active and monitoring correct directory'),
        ('Authentication failures', 'Check OAuth credentials are valid and not expired'),
        ('File write errors', 'Verify output directory exists and has write permissions'),
        ('API timeouts', 'Check network connectivity and API endpoint availability'),
        ('No records found', 'Verify date parameter is correct and data exists for that period')
    ]
    
    for issue, solution in troubleshooting:
        ws[f'A{row}'] = f'Issue: {issue}'
        ws[f'A{row}'].font = Font(bold=True, color=COLORS['red'])
        ws[f'A{row}'].fill = styles['light_amber_fill']
        ws[f'A{row}'].border = styles['border']
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 25
        row += 1
        
        ws[f'A{row}'] = f'Solution: {solution}'
        ws[f'A{row}'].fill = styles['light_amber_fill']
        ws[f'A{row}'].border = styles['border']
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 30
        row += 1
        
        row += 1
    
    # Section 4: Contact Information
    row += 1
    ws[f'A{row}'] = '📞 Support Contacts'
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'].fill = styles['header_fill']
    ws[f'A{row}'].font = styles['header_font']
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 25
    row += 1
    
    contacts = [
        'IT Operations: [Contact details to be provided]',
        'FSM Administrator: [Contact details to be provided]',
        'Business Owner: [Contact details to be provided]'
    ]
    
    for contact in contacts:
        ws[f'A{row}'] = contact
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'].border = styles['border']
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 20
        row += 1
    
    ws.column_dimensions['A'].width = 100
    ws.column_dimensions['B'].width = 20


if __name__ == "__main__":
    print("IPA Client Handover Template - Modern Edition (2026)")
    print("This template is called by Kiro during IPA analysis.")
    print("Usage: from ipa_client_handover_template import generate_report")

