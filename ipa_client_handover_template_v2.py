#!/usr/bin/env python3
"""
IPA Client Handover Documentation Template V2 - Comprehensive Edition (2026)

This template generates comprehensive client-facing IPA documentation that properly
extracts and displays ALL rich data from specialized subagent analysis.

Design: Modern color palette with comprehensive data extraction from nested structures.

Usage:
    from ipa_client_handover_template_v2 import generate_report
    generate_report(ipa_data)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from datetime import datetime
import os


# Modern Color Palette (2026 Professional)
COLORS = {
    'deep_blue': '1565C0',
    'green': '2E7D32',
    'amber': 'F57C00',
    'purple': '6A1B9A',
    'red': 'C62828',
    'medium_blue': '1E88E5',
    'light_blue': 'E3F2FD',
    'light_green': 'E8F5E9',
    'light_amber': 'FFF3E0',
    'light_purple': 'E1BEE7',
    'border_gray': 'BDBDBD',
    'white': 'FFFFFF'
}


def get_styles():
    """Return consistent style objects for modern design"""
    return {
        'header_fill': PatternFill(start_color=COLORS['deep_blue'], end_color=COLORS['deep_blue'], fill_type='solid'),
        'header_font': Font(color=COLORS['white'], bold=True, size=11),
        'subheader_fill': PatternFill(start_color=COLORS['medium_blue'], end_color=COLORS['medium_blue'], fill_type='solid'),
        'subheader_font': Font(color=COLORS['white'], bold=True, size=10),
        'success_fill': PatternFill(start_color=COLORS['green'], end_color=COLORS['green'], fill_type='solid'),
        'success_font': Font(color=COLORS['white'], bold=True),
        'warning_fill': PatternFill(start_color=COLORS['amber'], end_color=COLORS['amber'], fill_type='solid'),
        'warning_font': Font(color=COLORS['white'], bold=True),
        'error_fill': PatternFill(start_color=COLORS['red'], end_color=COLORS['red'], fill_type='solid'),
        'error_font': Font(color=COLORS['white'], bold=True),
        'api_fill': PatternFill(start_color=COLORS['purple'], end_color=COLORS['purple'], fill_type='solid'),
        'api_font': Font(color=COLORS['white'], bold=True),
        'light_blue_fill': PatternFill(start_color=COLORS['light_blue'], end_color=COLORS['light_blue'], fill_type='solid'),
        'light_green_fill': PatternFill(start_color=COLORS['light_green'], end_color=COLORS['light_green'], fill_type='solid'),
        'light_amber_fill': PatternFill(start_color=COLORS['light_amber'], end_color=COLORS['light_amber'], fill_type='solid'),
        'light_purple_fill': PatternFill(start_color=COLORS['light_purple'], end_color=COLORS['light_purple'], fill_type='solid'),
        'border': Border(
            left=Side(style='thin', color=COLORS['border_gray']),
            right=Side(style='thin', color=COLORS['border_gray']),
            top=Side(style='thin', color=COLORS['border_gray']),
            bottom=Side(style='thin', color=COLORS['border_gray'])
        )
    }


def generate_report(ipa_data):
    """
    Generate comprehensive client handover documentation Excel report.
    
    Args:
        ipa_data (dict): Dictionary containing all IPA analysis data from subagents
            Required keys:
            - client_name: str
            - process_group: str
            - business_requirements: dict (from business subagent)
            - workflow_analysis: dict (from workflow subagent)
            - configuration_guide: dict (from configuration subagent)
            - activity_guide: dict (from activity subagent)
            - production_validation: dict (from validation subagent)
    """
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Create all 6 comprehensive sheets
    create_executive_summary(wb, ipa_data)
    create_business_requirements(wb, ipa_data)
    create_workflow_and_approvals(wb, ipa_data)
    create_system_configuration(wb, ipa_data)
    create_activity_reference_guide(wb, ipa_data)
    create_production_validation(wb, ipa_data)
    
    # Save report
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = f"Client_Handover_Results/{ipa_data['client_name']}_{ipa_data['process_group']}_{timestamp}.xlsx"
    os.makedirs('Client_Handover_Results', exist_ok=True)
    wb.save(output_path)
    
    print(f"✓ Comprehensive client handover documentation generated: {output_path}")
    return output_path



def create_executive_summary(wb, ipa_data):
    """Create Executive Summary sheet with process overview and key features"""
    ws = wb.create_sheet('📊 Executive Summary')
    styles = get_styles()
    
    # Title
    ws['A1'] = f"📊 {ipa_data['client_name']} - {ipa_data['process_group']}"
    ws['A1'].font = Font(size=16, bold=True, color=COLORS['deep_blue'])
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:D1')
    ws.row_dimensions[1].height = 30
    
    # Date
    ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
    ws['A2'].font = Font(size=10, color=COLORS['border_gray'], italic=True)
    ws.merge_cells('A2:D2')
    
    row = 4
    
    # Key Features from objectives
    ws[f'A{row}'] = '✨ Key Features'
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'].fill = styles['success_fill']
    ws[f'A{row}'].font = styles['success_font']
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 25
    row += 1
    
    objectives = ipa_data.get('business_requirements', {}).get('objectives', [])
    for obj in objectives:
        # Handle both string and dict formats
        if isinstance(obj, dict):
            obj_text = obj.get('objective', str(obj))
        else:
            obj_text = str(obj)
        ws[f'A{row}'] = f"✓ {obj_text}"
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].fill = styles['light_green_fill']
        ws[f'A{row}'].border = styles['border']
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 30
        row += 1
    
    row += 1
    
    # Production Validation Summary
    test_summary = ipa_data.get('production_validation', {}).get('test_summary', {})
    if test_summary:
        ws[f'A{row}'] = '✅ Production Validation'
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].fill = styles['api_fill']
        ws[f'A{row}'].font = styles['api_font']
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 1
        
        validation_data = [
            ['Total Executions:', test_summary.get('total_executions', 'N/A')],
            ['Successful:', test_summary.get('successful', 'N/A')],
            ['Success Rate:', f"{test_summary.get('success_rate', 0)}%"],
            ['Test Date:', test_summary.get('test_date', 'N/A')],
            ['Environment:', test_summary.get('test_environment', 'N/A')]
        ]
        
        for label, value in validation_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws.merge_cells(f'B{row}:D{row}')
            for col in ['A', 'B', 'C', 'D']:
                cell = ws[f'{col}{row}']
                cell.border = styles['border']
                if row % 2 == 0:
                    cell.fill = styles['light_blue_fill']
            ws[f'A{row}'].font = Font(bold=True, color=COLORS['purple'])
            row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25



def create_business_requirements(wb, ipa_data):
    """Create comprehensive Business Requirements sheet extracting ALL subagent data"""
    ws = wb.create_sheet('📋 Business Requirements')
    styles = get_styles()
    
    # Title
    ws['A1'] = '📋 Business Requirements'
    ws['A1'].font = Font(size=16, bold=True, color=COLORS['deep_blue'])
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:E1')
    ws.row_dimensions[1].height = 30
    
    row = 3
    
    # Requirements Section
    requirements = ipa_data.get('business_requirements', {}).get('requirements', [])
    if requirements:
        ws[f'A{row}'] = '🎯 Requirements'
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'].fill = styles['header_fill']
        ws[f'A{row}'].font = styles['header_font']
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 1
        
        # Headers
        headers = ['ID', 'Category', 'Description', 'Priority', 'Source']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = styles['subheader_fill']
            cell.font = styles['subheader_font']
            cell.border = styles['border']
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 20
        row += 1
        
        # Requirements data
        for req in requirements:
            ws.cell(row=row, column=1, value=req.get('id', 'N/A'))
            ws.cell(row=row, column=2, value=req.get('category', 'N/A'))
            ws.cell(row=row, column=3, value=req.get('description', 'N/A'))
            ws.cell(row=row, column=4, value=req.get('priority', 'Medium'))
            ws.cell(row=row, column=5, value=req.get('source', 'N/A'))
            
            for col_idx in range(1, 6):
                cell = ws.cell(row=row, column=col_idx)
                cell.border = styles['border']
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if row % 2 == 0:
                    cell.fill = styles['light_blue_fill']
            
            # Color code priority
            priority = req.get('priority', 'Medium')
            if priority == 'High':
                ws.cell(row=row, column=4).font = Font(bold=True, color=COLORS['red'])
            elif priority == 'Medium':
                ws.cell(row=row, column=4).font = Font(bold=True, color=COLORS['amber'])
            
            ws.row_dimensions[row].height = 40
            row += 1
        
        row += 1
    
    # Stakeholders Section
    stakeholders = ipa_data.get('business_requirements', {}).get('stakeholders', [])
    if stakeholders:
        ws[f'A{row}'] = '👥 Stakeholders'
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'].fill = styles['success_fill']
        ws[f'A{row}'].font = styles['success_font']
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 1
        
        # Headers
        headers = ['Role', 'Interest', 'Impact']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = styles['subheader_fill']
            cell.font = styles['subheader_font']
            cell.border = styles['border']
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 20
        row += 1
        
        # Stakeholder data
        for stakeholder in stakeholders:
            ws.cell(row=row, column=1, value=stakeholder.get('role', 'N/A'))
            ws.cell(row=row, column=2, value=stakeholder.get('interest', 'N/A'))
            ws.cell(row=row, column=3, value=stakeholder.get('impact', 'N/A'))
            
            for col_idx in range(1, 4):
                cell = ws.cell(row=row, column=col_idx)
                cell.border = styles['border']
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if row % 2 == 0:
                    cell.fill = styles['light_green_fill']
            
            ws.row_dimensions[row].height = 30
            row += 1
        
        row += 1
    
    # Scope Section
    scope = ipa_data.get('business_requirements', {}).get('scope', {})
    if scope:
        ws[f'A{row}'] = '📦 Scope'
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'].fill = styles['warning_fill']
        ws[f'A{row}'].font = styles['warning_font']
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 1
        
        # In Scope
        in_scope = scope.get('in_scope', [])
        if in_scope:
            ws[f'A{row}'] = '✅ In Scope'
            ws[f'A{row}'].font = Font(bold=True, color=COLORS['green'])
            ws[f'A{row}'].border = styles['border']
            ws.row_dimensions[row].height = 20
            row += 1
            
            for item in in_scope:
                ws[f'A{row}'] = f"  • {item}"
                ws.merge_cells(f'A{row}:E{row}')
                ws[f'A{row}'].border = styles['border']
                ws[f'A{row}'].alignment = Alignment(wrap_text=True)
                ws.row_dimensions[row].height = 25
                row += 1
        
        # Out of Scope
        out_of_scope = scope.get('out_of_scope', [])
        if out_of_scope:
            ws[f'A{row}'] = '❌ Out of Scope'
            ws[f'A{row}'].font = Font(bold=True, color=COLORS['red'])
            ws[f'A{row}'].border = styles['border']
            ws.row_dimensions[row].height = 20
            row += 1
            
            for item in out_of_scope:
                ws[f'A{row}'] = f"  • {item}"
                ws.merge_cells(f'A{row}:E{row}')
                ws[f'A{row}'].border = styles['border']
                ws[f'A{row}'].alignment = Alignment(wrap_text=True)
                ws.row_dimensions[row].height = 25
                row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20



def create_workflow_and_approvals(wb, ipa_data):
    """Create Workflow & Approval Paths sheet"""
    ws = wb.create_sheet('🔄 Workflow & Approvals')
    styles = get_styles()
    
    ws['A1'] = '🔄 Workflow & Approval Paths'
    ws['A1'].font = Font(size=16, bold=True, color=COLORS['deep_blue'])
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:D1')
    ws.row_dimensions[1].height = 30
    
    row = 3
    
    # Workflow Steps
    # Try nested structure first, then flat structure
    workflow_steps = ipa_data.get('workflow_analysis', {}).get('workflow_steps', [])
    if not workflow_steps:
        workflow_steps = ipa_data.get('workflow_steps', [])
    if workflow_steps:
        ws[f'A{row}'] = '📝 Workflow Steps'
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].fill = styles['header_fill']
        ws[f'A{row}'].font = styles['header_font']
        ws.row_dimensions[row].height = 25
        row += 1
        
        for idx, step in enumerate(workflow_steps, start=1):
            # Handle both old format (step/activity/description) and new format (type/label)
            if 'label' in step:
                # New format from workflow analyzer
                label = step.get('label', 'N/A')
                step_type = step.get('type', 'process')
            else:
                # Old format (fallback)
                step_num = step.get('step', idx)
                activity = step.get('activity', 'N/A')
                step_type = step.get('type', 'process')
                description = step.get('description', '')
                business_purpose = step.get('business_purpose', '')
                
                # Build step label
                label = f"Step {step_num}: {activity}"
                if description:
                    label += f"\n{description}"
                if business_purpose:
                    label += f"\n→ {business_purpose}"
            
            ws[f'A{row}'] = label
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'].border = styles['border']
            ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            
            # Color code by type
            if step_type == 'start':
                ws[f'A{row}'].fill = styles['light_green_fill']
            elif step_type == 'end':
                ws[f'A{row}'].fill = styles['light_green_fill']
            elif step_type == 'decision':
                ws[f'A{row}'].fill = styles['light_amber_fill']
            elif step_type in ['api', 'approval']:
                ws[f'A{row}'].fill = styles['light_purple_fill']
            elif step_type == 'file':
                ws[f'A{row}'].fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
            else:
                ws[f'A{row}'].fill = styles['light_blue_fill']
            
            ws.row_dimensions[row].height = 60
            row += 1
        
        row += 1
    else:
        # Show message if no workflow steps
        ws[f'A{row}'] = 'N/A'
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].border = styles['border']
        ws.row_dimensions[row].height = 25
        row += 2
    
    # Integrations
    # Try nested structure first, then flat structure
    integrations = ipa_data.get('workflow_analysis', {}).get('integrations', [])
    if not integrations:
        integrations = ipa_data.get('integrations', [])
    if integrations:
        ws[f'A{row}'] = '🌐 Integrations'
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].fill = styles['api_fill']
        ws[f'A{row}'].font = styles['api_font']
        ws.row_dimensions[row].height = 25
        row += 1
        
        for integration in integrations:
            int_type = integration.get('type', 'N/A')
            description = integration.get('description', 'N/A')
            when = integration.get('when', 'N/A')
            
            ws[f'A{row}'] = f"{int_type}: {description}"
            ws[f'D{row}'] = when
            ws.merge_cells(f'A{row}:C{row}')
            
            for col in ['A', 'B', 'C', 'D']:
                cell = ws[f'{col}{row}']
                cell.border = styles['border']
                cell.alignment = Alignment(wrap_text=True)
                if row % 2 == 0:
                    cell.fill = styles['light_purple_fill']
            
            ws.row_dimensions[row].height = 30
            row += 1
    
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30


def create_system_configuration(wb, ipa_data):
    """Create comprehensive System Configuration sheet"""
    ws = wb.create_sheet('⚙️ System Configuration')
    styles = get_styles()
    
    ws['A1'] = '⚙️ System Configuration'
    ws['A1'].font = Font(size=16, bold=True, color=COLORS['purple'])
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:E1')
    ws.row_dimensions[1].height = 30
    
    row = 3
    
    # Configuration Variables
    # Try nested structure first, then flat structure
    config_vars = ipa_data.get('configuration_guide', {}).get('config_variables', [])
    if not config_vars:
        config_vars = ipa_data.get('config_variables', [])
    if config_vars:
        ws[f'A{row}'] = '📝 Configuration Variables'
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'].fill = styles['header_fill']
        ws[f'A{row}'].font = styles['header_font']
        ws.row_dimensions[row].height = 25
        row += 1
        
        # Headers
        headers = ['Variable Name', 'Type', 'Location', 'Description', 'How to Modify']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = styles['subheader_fill']
            cell.font = styles['subheader_font']
            cell.border = styles['border']
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 20
        row += 1
        
        # Config data
        for var in config_vars:
            ws.cell(row=row, column=1, value=var.get('name', 'N/A'))
            ws.cell(row=row, column=2, value=var.get('type', 'N/A'))
            ws.cell(row=row, column=3, value=var.get('location', 'N/A'))
            ws.cell(row=row, column=4, value=var.get('description', 'N/A'))
            ws.cell(row=row, column=5, value=var.get('how_to_modify', 'N/A'))
            
            for col_idx in range(1, 6):
                cell = ws.cell(row=row, column=col_idx)
                cell.border = styles['border']
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if row % 2 == 0:
                    cell.fill = styles['light_blue_fill']
            
            ws.row_dimensions[row].height = 60
            row += 1
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 50


def create_activity_reference_guide(wb, ipa_data):
    """Create comprehensive Activity Reference Guide sheet"""
    ws = wb.create_sheet('📚 Activity Reference')
    styles = get_styles()
    
    ws['A1'] = '📚 Activity Reference Guide'
    ws['A1'].font = Font(size=16, bold=True, color=COLORS['deep_blue'])
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = 30
    
    row = 3
    
    # Activities
    # Try nested structure first, then flat structure
    activities = ipa_data.get('activity_guide', {}).get('activities', [])
    if not activities:
        # Try maintenance_guide.activity_guide structure
        maintenance_guide = ipa_data.get('maintenance_guide', {})
        if isinstance(maintenance_guide, dict):
            activity_guide = maintenance_guide.get('activity_guide', {})
            if isinstance(activity_guide, dict):
                activities = activity_guide.get('activities', [])
        # Try flat structure
        if not activities:
            activities = ipa_data.get('activities', [])
    if activities:
        ws[f'A{row}'] = '🔧 Activities'
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'].fill = styles['header_fill']
        ws[f'A{row}'].font = styles['header_font']
        ws.row_dimensions[row].height = 25
        row += 1
        
        # Headers
        headers = ['ID', 'Caption', 'Description', 'When It Runs', 'What It Does', 'Maintenance Notes']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = styles['subheader_fill']
            cell.font = styles['subheader_font']
            cell.border = styles['border']
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 20
        row += 1
        
        # Activity data
        for activity in activities:
            ws.cell(row=row, column=1, value=activity.get('id', 'N/A'))
            ws.cell(row=row, column=2, value=activity.get('caption', 'N/A'))
            ws.cell(row=row, column=3, value=activity.get('description', 'N/A'))
            ws.cell(row=row, column=4, value=activity.get('when_it_runs', 'N/A'))
            ws.cell(row=row, column=5, value=activity.get('what_it_does', 'N/A'))
            ws.cell(row=row, column=6, value=activity.get('maintenance_notes', 'N/A'))
            
            for col_idx in range(1, 7):
                cell = ws.cell(row=row, column=col_idx)
                cell.border = styles['border']
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if row % 2 == 0:
                    cell.fill = styles['light_blue_fill']
            
            # Highlight configurable activities
            if activity.get('configurable', False):
                ws.cell(row=row, column=6).fill = styles['light_green_fill']
            
            ws.row_dimensions[row].height = 50
            row += 1
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 40


def create_production_validation(wb, ipa_data):
    """Create comprehensive Production Validation sheet"""
    ws = wb.create_sheet('✅ Production Validation')
    styles = get_styles()
    
    ws['A1'] = '✅ Production Validation'
    ws['A1'].font = Font(size=16, bold=True, color=COLORS['green'])
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:D1')
    ws.row_dimensions[1].height = 30
    
    row = 3
    
    # Test Summary
    test_summary = ipa_data.get('production_validation', {}).get('test_summary', {})
    if test_summary:
        ws[f'A{row}'] = '📊 Test Summary'
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].fill = styles['success_fill']
        ws[f'A{row}'].font = styles['success_font']
        ws.row_dimensions[row].height = 25
        row += 1
        
        summary_data = [
            ['Work Unit', test_summary.get('work_unit_number', test_summary.get('work_unit', 'N/A'))],
            ['Total Executions', test_summary.get('total_executions', 'N/A')],
            ['Successful', test_summary.get('successful', 'N/A')],
            ['Failed', test_summary.get('failed', 'N/A')],
            ['Success Rate', f"{test_summary.get('success_rate', 0)}%"],
            ['Test Date', test_summary.get('test_date', 'N/A')],
            ['Environment', test_summary.get('test_environment', 'N/A')]
        ]
        
        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws.merge_cells(f'B{row}:D{row}')
            for col in ['A', 'B', 'C', 'D']:
                cell = ws[f'{col}{row}']
                cell.border = styles['border']
                if row % 2 == 0:
                    cell.fill = styles['light_green_fill']
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        row += 1
    
    # Performance
    performance = ipa_data.get('production_validation', {}).get('performance', {})
    if performance:
        ws[f'A{row}'] = '⚡ Performance'
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].fill = styles['warning_fill']
        ws[f'A{row}'].font = styles['warning_font']
        ws.row_dimensions[row].height = 25
        row += 1
        
        perf_data = [
            ['Average Duration', performance.get('avg_duration', 'N/A')],
            ['Min Duration', performance.get('min_duration', 'N/A')],
            ['Max Duration', performance.get('max_duration', 'N/A')],
            ['Performance Rating', performance.get('performance_rating', 'N/A')],
            ['Notes', performance.get('notes', 'N/A')]
        ]
        
        for label, value in perf_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws.merge_cells(f'B{row}:D{row}')
            for col in ['A', 'B', 'C', 'D']:
                cell = ws[f'{col}{row}']
                cell.border = styles['border']
                cell.alignment = Alignment(wrap_text=True)
                if row % 2 == 0:
                    cell.fill = styles['light_amber_fill']
            ws[f'A{row}'].font = Font(bold=True)
            ws.row_dimensions[row].height = 30
            row += 1
        
        row += 1
    
    # Production Readiness
    readiness = ipa_data.get('production_validation', {}).get('production_readiness', {})
    if readiness:
        ws[f'A{row}'] = '🎯 Production Readiness'
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].fill = styles['api_fill']
        ws[f'A{row}'].font = styles['api_font']
        ws.row_dimensions[row].height = 25
        row += 1
        
        ws[f'A{row}'] = 'Ready'
        ws[f'B{row}'] = '✅ Yes' if readiness.get('ready', False) else '❌ No'
        ws.merge_cells(f'B{row}:D{row}')
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].border = styles['border']
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = 'Confidence Level'
        ws[f'B{row}'] = readiness.get('confidence_level', 'N/A')
        ws.merge_cells(f'B{row}:D{row}')
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].border = styles['border']
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        # Evidence
        evidence = readiness.get('evidence', [])
        if evidence:
            ws[f'A{row}'] = 'Evidence'
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = styles['border']
            row += 1
            
            for item in evidence:
                ws[f'A{row}'] = f"  ✓ {item}"
                ws.merge_cells(f'A{row}:D{row}')
                ws[f'A{row}'].border = styles['border']
                ws[f'A{row}'].alignment = Alignment(wrap_text=True)
                ws.row_dimensions[row].height = 25
                row += 1
        
        # Recommendations
        recommendations = readiness.get('recommendations', [])
        if recommendations:
            row += 1
            ws[f'A{row}'] = 'Recommendations'
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = styles['border']
            row += 1
            
            for item in recommendations:
                ws[f'A{row}'] = f"  → {item}"
                ws.merge_cells(f'A{row}:D{row}')
                ws[f'A{row}'].border = styles['border']
                ws[f'A{row}'].alignment = Alignment(wrap_text=True)
                ws.row_dimensions[row].height = 25
                row += 1
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30


if __name__ == "__main__":
    print("IPA Client Handover Template V2 - Comprehensive Edition")
    print("Import this module and call generate_report(ipa_data)")
