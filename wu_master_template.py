#!/usr/bin/env python3
"""
================================================================================
WORK UNIT ANALYSIS MASTER TEMPLATE
================================================================================

PURPOSE:
    This template generates professional Excel reports for Work Unit (WU) analysis.
    It is a REPORT GENERATOR ONLY - it does NOT perform analysis.
    
    Kiro (the AI assistant) performs the analysis by:
    1. Reading the work unit log file
    2. Parsing activities, errors, JavaScript code, and metrics
    3. Building the wu_data dictionary with analyzed data
    4. Calling generate_report(wu_data) to create the Excel report

USAGE:
    From Kiro (recommended):
        1. Kiro reads and analyzes the WU log file
        2. Kiro builds the wu_data dictionary
        3. Kiro calls: generate_report(wu_data)
    
    From command line (with pre-analyzed JSON):
        python wu_master_template.py <analysis_data.json>

OUTPUT:
    Excel workbook with 9 sheets:
    - Summary Dashboard: Executive overview with charts
    - Activity Timeline: Chronological activity execution
    - Performance Metrics: Duration, memory, CPU metrics
    - Error Analysis: Detected errors and their impact
    - JavaScript Review: ES5 compliance issues (10 columns)
    - SQL Review: Data Fabric/Compass SQL query analysis (10 columns)
    - Memory Analysis: Memory usage by activity
    - Recommendations: Actionable fixes (10 columns)
    - Technical Analysis: Deep dive architecture review

DATA STRUCTURE (wu_data dictionary):
    {
        'work_unit_id': str,           # e.g., "636228"
        'process_name': str,           # e.g., "CISOutboundIntegration"
        'status': str,                 # "SUCCESS" or "FAILED"
        
        'info_data': [                 # Dashboard metrics rows
            ["Metric Name", "Value", "Status", "Rating"],
            ...
        ],
        
        'chart_data': {
            'status': [("Success", 1), ("Failed", 0)],
            'memory_by_activity': [("Activity1", 100.5), ...]
        },
        
        'activities': [                # Activity timeline
            {
                'name': str,
                'type': str,           # START, LM, ASSGN, BRANCH, etc.
                'start_time': str,
                'end_time': str,
                'duration': str,
                'status': str          # Completed, Error, etc.
            },
            ...
        ],
        
        'metrics': {                   # Performance metrics
            'duration_ms': int,
            'memory_mib': float,
            'cpu_time_ms': int,
            'user_time_ms': int
        },
        
        'errors': [                    # Error analysis
            {
                'activity': str,
                'type': str,
                'message': str,
                'severity': str,       # Critical, High, Medium, Low
                'impact': str
            },
            ...
        ],
        
        'js_issues': [                 # JavaScript review (10 columns)
            [work_unit_id, activity, block, issue_type, severity, 
             line, code_snippet, root_cause, specific_fix_es5, prevention],
            ...
        ],
        
        'sql_issues': [                # SQL review (10 columns) - for Data Fabric queries
            [work_unit_id, activity, query_type, sql_snippet, severity,
             issue, root_cause, recommendation, compass_compatible, performance_impact],
            ...
        ],
        
        'memory_analysis': [           # Memory by activity
            [work_unit_id, activity, type, memory_mib, duration_ms, efficiency, rating],
            ...
        ],
        
        'recommendations': [           # Recommendations (10 columns)
            [priority, category, issue_description, affected_wus, root_cause,
             specific_fix, code_example_es5, testing_steps, effort, impact],
            ...
        ],
        
        'architecture_analysis': [     # Technical analysis - architecture
            ["Aspect", "Value", "Assessment"],
            ...
        ],
        
        'js_quality_analysis': [       # Technical analysis - JS quality
            ["Metric", "Score", "Assessment"],
            ...
        ],
        
        'business_impact': [           # Technical analysis - business impact
            ["Factor", "Value", "Assessment"],
            ...
        ]
    }

ES5 JAVASCRIPT REQUIREMENTS:
    IPA uses ES5 JavaScript engine. All fix recommendations must use ES5 syntax:
    - Use 'var' (NOT 'let' or 'const')
    - Use 'typeof varName !== "undefined"' for null checks
    - Use 'function()' (NOT arrow functions '=>')
    - Use string concatenation with '+' (NOT template literals)
    - Use 'Array.prototype.slice.call()' (NOT spread operator '...')

REPORT LOCATION:
    Reports are saved to: Performance_Results/WU_<ProcessName>_Report_YYYYMMDD.xlsx

DEPENDENCIES:
    - openpyxl (Excel generation)
    - json (data loading from file)

AUTHOR: Van Anthony Silleza - FSM Technical Consultant
        Kiro - AI Assistant
VERSION: 1.0
================================================================================
"""

import os
import sys
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import PieChart, BarChart, Reference

# Style constants
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)



def create_summary_dashboard(wb, wu_data):
    """Create the Summary Dashboard sheet with executive overview and charts."""
    ws = wb.active
    ws.title = "Summary Dashboard"
    
    # Title
    ws.merge_cells('A1:H1')
    ws.cell(row=1, column=1, value="Work Unit " + str(wu_data['work_unit_id']) + " - Executive Dashboard")
    ws.cell(row=1, column=1).font = Font(size=18, bold=True, color="0066CC")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    
    # Process Information section
    ws.merge_cells('A3:D3')
    ws.cell(row=3, column=1, value="Process Information")
    ws.cell(row=3, column=1).font = Font(bold=True, size=14, color="0066CC")
    ws.cell(row=3, column=1).fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
    
    # Headers
    for col, header in enumerate(["Metric", "Value", "Status", "Rating"], start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center')
    
    # Info data rows
    for row_idx, row_data in enumerate(wu_data.get('info_data', []), start=5):
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = BORDER
            if col == 1:
                cell.font = Font(bold=True)
            if col in [3, 4]:
                v = str(value)
                if "FAILED" in v or "Critical" in v or "Red" in v:
                    cell.fill = RED_FILL
                    cell.font = WHITE_FONT
                elif "Excellent" in v or "Green" in v or "SUCCESS" in v or "Good" in v:
                    cell.fill = GREEN_FILL
                elif "Warning" in v or "Yellow" in v:
                    cell.fill = YELLOW_FILL
    
    # Performance Overview section
    ws.merge_cells('F3:K3')
    ws.cell(row=3, column=6, value="Performance Overview")
    ws.cell(row=3, column=6).font = Font(bold=True, size=14, color="0066CC")
    
    # Chart data
    chart_data = wu_data.get('chart_data', {})
    for i, item in enumerate(chart_data.get('status', []), start=5):
        ws.cell(row=i, column=10, value=item[0])
        ws.cell(row=i, column=11, value=item[1])
    for i, item in enumerate(chart_data.get('memory_by_activity', [])[:5], start=8):
        ws.cell(row=i, column=10, value=item[0][:15])
        ws.cell(row=i, column=11, value=item[1])
    
    # Pie chart for status
    status_data = chart_data.get('status', [])
    if status_data:
        pie = PieChart()
        pie.title = "Process Status"
        pie.add_data(Reference(ws, min_col=11, min_row=5, max_row=5+len(status_data)-1))
        pie.set_categories(Reference(ws, min_col=10, min_row=5, max_row=5+len(status_data)-1))
        pie.width = 6
        pie.height = 5
        ws.add_chart(pie, "F4")
    
    # Bar chart for memory
    mem_data = chart_data.get('memory_by_activity', [])
    if mem_data:
        bar = BarChart()
        bar.title = "Memory by Activity"
        bar.add_data(Reference(ws, min_col=11, min_row=8, max_row=8+min(len(mem_data),5)-1))
        bar.set_categories(Reference(ws, min_col=10, min_row=8, max_row=8+min(len(mem_data),5)-1))
        bar.width = 6
        bar.height = 5
        ws.add_chart(bar, "I4")
    
    # Column widths
    for col, w in enumerate([18,20,15,12,15,12,12,12,12,8,8], start=1):
        ws.column_dimensions[chr(64+col)].width = w
    return ws


def create_activity_timeline(wb, wu_data):
    """Create the Activity Timeline sheet with chronological activity execution."""
    ws = wb.create_sheet("Activity Timeline")
    for col, h in enumerate(["Work Unit ID","Activity","Type","Start","End","Duration","Status"], start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
    for row, a in enumerate(wu_data.get('activities', []), start=2):
        data = [wu_data['work_unit_id'], a.get('name',''), a.get('type',''), 
                a.get('start_time',''), a.get('end_time','N/A'), a.get('duration','N/A'), a.get('status','Completed')]
        for col, v in enumerate(data, start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = BORDER
            if col == 7 and v in ["Error","Failed"]:
                cell.fill = RED_FILL
                cell.font = WHITE_FONT
    for col in range(1,8):
        ws.column_dimensions[chr(64+col)].width = 18
    return ws


def create_performance_metrics(wb, wu_data):
    """Create the Performance Metrics sheet with duration, memory, CPU metrics."""
    ws = wb.create_sheet("Performance Metrics")
    for col, h in enumerate(["Work Unit ID","Process Name","Duration (ms)","Memory (MiB)","CPU Time (ms)","User Time (ms)","Status"], start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
    m = wu_data.get('metrics', {})
    data = [wu_data['work_unit_id'], wu_data.get('process_name',''), m.get('duration_ms',0),
            m.get('memory_mib',0), m.get('cpu_time_ms',0), m.get('user_time_ms',0), wu_data.get('status','Unknown')]
    for col, v in enumerate(data, start=1):
        cell = ws.cell(row=2, column=col, value=v)
        cell.border = BORDER
        if col == 4:
            mem = float(v) if v else 0
            cell.fill = GREEN_FILL if mem < 70000 else YELLOW_FILL if mem < 90000 else RED_FILL
            if mem >= 90000:
                cell.font = WHITE_FONT
        if col == 7 and v == "FAILED":
            cell.fill = RED_FILL
            cell.font = WHITE_FONT
    for col in range(1,8):
        ws.column_dimensions[chr(64+col)].width = 18
    return ws


def create_error_analysis(wb, wu_data):
    """Create the Error Analysis sheet with detected errors and their impact."""
    ws = wb.create_sheet("Error Analysis")
    for col, h in enumerate(["Work Unit ID","Activity","Error Type","Error Message","Severity","Impact"], start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
    errors = wu_data.get('errors', [])
    if errors:
        for row, e in enumerate(errors, start=2):
            data = [wu_data['work_unit_id'], e.get('activity','N/A'), e.get('type',''),
                    e.get('message','')[:100], e.get('severity','High'), e.get('impact','Process Failure')]
            for col, v in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.border = BORDER
                if col in [5,6] and "Critical" in str(v):
                    cell.fill = RED_FILL
                    cell.font = WHITE_FONT
    else:
        for col, v in enumerate([wu_data['work_unit_id'],"N/A","None","No errors detected","N/A","N/A"], start=1):
            ws.cell(row=2, column=col, value=v).border = BORDER
    for col in range(1,7):
        ws.column_dimensions[chr(64+col)].width = 20
    return ws



def create_javascript_review(wb, wu_data):
    """Create the JavaScript Review sheet with ES5 compliance issues (10 columns)."""
    ws = wb.create_sheet("JavaScript Review")
    headers = ["Work Unit ID","Activity","Block","Issue Type","Severity","Line","Code Snippet","Root Cause","Specific Fix (ES5)","Prevention"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(wrap_text=True)
    js_issues = wu_data.get('js_issues', [])
    if js_issues:
        for row, issue in enumerate(js_issues, start=2):
            for col, v in enumerate(issue, start=1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.border = BORDER
                cell.alignment = Alignment(wrap_text=True)
                if col == 5:
                    if v == "Critical":
                        cell.fill = RED_FILL
                        cell.font = WHITE_FONT
                    elif v == "High":
                        cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                    elif v == "Medium":
                        cell.fill = YELLOW_FILL
                    elif v == "Low":
                        cell.fill = GREEN_FILL
            ws.row_dimensions[row].height = 60
    else:
        for col, v in enumerate([wu_data['work_unit_id'],"N/A","N/A","No Issues","N/A","N/A","N/A","N/A","N/A","N/A"], start=1):
            ws.cell(row=2, column=col, value=v).border = BORDER
    for col, w in enumerate([12,15,8,20,12,8,30,35,40,30], start=1):
        ws.column_dimensions[chr(64+col)].width = w
    return ws


def create_sql_review(wb, wu_data):
    """Create the SQL Review sheet for Data Fabric/Compass API queries (10 columns)."""
    ws = wb.create_sheet("SQL Review")
    headers = ["Work Unit ID", "Activity", "Query Type", "SQL Snippet", "Severity", 
               "Issue", "Root Cause", "Recommendation", "Compass Compatible", "Performance Impact"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(wrap_text=True)
    
    sql_issues = wu_data.get('sql_issues', [])
    if sql_issues:
        for row, issue in enumerate(sql_issues, start=2):
            for col, v in enumerate(issue, start=1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.border = BORDER
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                # Color code severity
                if col == 5:
                    sev = str(v).upper()
                    if sev == "CRITICAL":
                        cell.fill = RED_FILL
                        cell.font = WHITE_FONT
                    elif sev == "HIGH":
                        cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                    elif sev == "MEDIUM":
                        cell.fill = YELLOW_FILL
                    elif sev in ["LOW", "INFO"]:
                        cell.fill = GREEN_FILL
                # Color code Compass compatibility
                if col == 9:
                    compat = str(v).upper()
                    if compat in ["YES", "TRUE", "COMPATIBLE"]:
                        cell.fill = GREEN_FILL
                    elif compat in ["NO", "FALSE", "INCOMPATIBLE"]:
                        cell.fill = RED_FILL
                        cell.font = WHITE_FONT
                    elif compat in ["PARTIAL", "REVIEW"]:
                        cell.fill = YELLOW_FILL
            ws.row_dimensions[row].height = 80
    else:
        # No SQL queries found
        for col, v in enumerate([wu_data['work_unit_id'], "N/A", "None", "No SQL queries detected in WEBRN activities", 
                                  "Info", "N/A", "N/A", "N/A", "N/A", "N/A"], start=1):
            cell = ws.cell(row=2, column=col, value=v)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True)
    
    # Column widths
    for col, w in enumerate([12, 15, 12, 50, 10, 30, 35, 40, 15, 20], start=1):
        ws.column_dimensions[chr(64+col)].width = w
    return ws


def create_memory_analysis(wb, wu_data):
    """Create the Memory Analysis sheet with memory usage by activity."""
    ws = wb.create_sheet("Memory Analysis")
    for col, h in enumerate(["Work Unit ID","Activity","Type","Memory (MiB)","Duration (ms)","Efficiency","Rating"], start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
    for row, item in enumerate(wu_data.get('memory_analysis', []), start=2):
        for col, v in enumerate(item, start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = BORDER
            if col == 4:
                try:
                    mem = float(v)
                    cell.fill = GREEN_FILL if mem < 70 else YELLOW_FILL if mem < 500 else RED_FILL
                    if mem >= 500:
                        cell.font = WHITE_FONT
                except:
                    pass
    for col in range(1,8):
        ws.column_dimensions[chr(64+col)].width = 15
    return ws


def create_recommendations(wb, wu_data):
    """Create the Recommendations sheet with actionable fixes (10 columns)."""
    ws = wb.create_sheet("Recommendations")
    headers = ["Priority","Category","Issue Description","Affected WUs","Root Cause Analysis","Specific Fix Required","Code Example (ES5)","Testing Steps","Effort","Impact"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(wrap_text=True)
    recs = wu_data.get('recommendations', [])
    if not recs:
        recs = [["LOW","General","No critical issues found",wu_data['work_unit_id'],"Process executed without critical errors","Continue monitoring","N/A","Regular monitoring","N/A","N/A"]]
    for row, rec in enumerate(recs, start=2):
        for col, v in enumerate(rec, start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if col == 1:
                if v == "CRITICAL":
                    cell.fill = RED_FILL
                    cell.font = WHITE_FONT
                elif v == "HIGH":
                    cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                elif v == "MEDIUM":
                    cell.fill = YELLOW_FILL
                elif v == "LOW":
                    cell.fill = GREEN_FILL
        ws.row_dimensions[row].height = 100
    for col, w in enumerate([12,15,25,12,35,30,45,30,10,15], start=1):
        ws.column_dimensions[chr(64+col)].width = w
    return ws


def create_technical_analysis(wb, wu_data):
    """Create the Technical Analysis sheet with deep dive architecture review."""
    ws = wb.create_sheet("Technical Analysis")
    ws.cell(row=1, column=1, value="Technical Deep Dive - Work Unit " + str(wu_data['work_unit_id']))
    ws.cell(row=1, column=1).font = Font(size=16, bold=True, color="0066CC")
    ws.merge_cells('A1:H1')
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    
    # Process Architecture Analysis
    ws.cell(row=3, column=1, value="Process Architecture Analysis").font = Font(bold=True, size=14, color="0066CC")
    for i, item in enumerate(wu_data.get('architecture_analysis', []), start=4):
        ws.cell(row=i, column=1, value=item[0]).font = Font(bold=True)
        ws.cell(row=i, column=2, value=item[1])
        ws.cell(row=i, column=3, value=item[2])
        for col in range(1,4):
            ws.cell(row=i, column=col).border = BORDER
        a = str(item[2])
        if "Critical" in a or "Failed" in a:
            ws.cell(row=i, column=3).fill = RED_FILL
            ws.cell(row=i, column=3).font = WHITE_FONT
        elif "Warning" in a or "Review" in a:
            ws.cell(row=i, column=3).fill = YELLOW_FILL
        elif "Compliant" in a or "Good" in a or "Excellent" in a:
            ws.cell(row=i, column=3).fill = GREEN_FILL
    
    # JavaScript Code Quality (ES5)
    js_row = len(wu_data.get('architecture_analysis', [])) + 6
    ws.cell(row=js_row, column=1, value="JavaScript Code Quality (ES5)").font = Font(bold=True, size=14, color="0066CC")
    for i, item in enumerate(wu_data.get('js_quality_analysis', []), start=js_row+1):
        ws.cell(row=i, column=1, value=item[0]).font = Font(bold=True)
        ws.cell(row=i, column=2, value=item[1])
        ws.cell(row=i, column=3, value=item[2])
        for col in range(1,4):
            ws.cell(row=i, column=col).border = BORDER
        a = str(item[2])
        if "Critical" in a or "Issues" in a:
            ws.cell(row=i, column=3).fill = RED_FILL
            ws.cell(row=i, column=3).font = WHITE_FONT
        elif "Review" in a or "Warning" in a:
            ws.cell(row=i, column=3).fill = YELLOW_FILL
        elif "Compliant" in a or "Good" in a or "None" in a:
            ws.cell(row=i, column=3).fill = GREEN_FILL
    
    # Business Impact Analysis
    impact_row = js_row + len(wu_data.get('js_quality_analysis', [])) + 3
    ws.cell(row=impact_row, column=1, value="Business Impact Analysis").font = Font(bold=True, size=14, color="0066CC")
    for i, item in enumerate(wu_data.get('business_impact', []), start=impact_row+1):
        ws.cell(row=i, column=1, value=item[0]).font = Font(bold=True)
        ws.cell(row=i, column=2, value=item[1])
        ws.cell(row=i, column=3, value=item[2])
        for col in range(1,4):
            ws.cell(row=i, column=col).border = BORDER
        a = str(item[2])
        if "Critical" in a or "Failed" in a:
            ws.cell(row=i, column=3).fill = RED_FILL
            ws.cell(row=i, column=3).font = WHITE_FONT
        elif "Review" in a or "Risk" in a:
            ws.cell(row=i, column=3).fill = YELLOW_FILL
        elif "Good" in a or "Success" in a or "Normal" in a:
            ws.cell(row=i, column=3).fill = GREEN_FILL
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 50
    return ws


def generate_report(wu_data, output_dir="Performance_Results"):
    """
    Generate Excel report from Kiro's pre-analyzed data.
    
    Args:
        wu_data: Dictionary containing all analyzed work unit data
        output_dir: Directory to save the report (default: Performance_Results)
    
    Returns:
        str: Path to the generated report file
    """
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    wb = Workbook()
    create_summary_dashboard(wb, wu_data)
    create_activity_timeline(wb, wu_data)
    create_performance_metrics(wb, wu_data)
    create_error_analysis(wb, wu_data)
    create_javascript_review(wb, wu_data)
    create_sql_review(wb, wu_data)
    create_memory_analysis(wb, wu_data)
    create_technical_analysis(wb, wu_data)
    create_recommendations(wb, wu_data)
    
    proc = str(wu_data.get('process_name', 'Unknown')).replace(' ', '_').replace('/', '_')[:30]
    filename = output_dir + "/WU_" + proc + "_Report_" + datetime.now().strftime('%Y%m%d') + ".xlsx"
    wb.save(filename)
    return filename


def main():
    """Command-line entry point for generating reports from JSON data files."""
    if len(sys.argv) < 2:
        print("Usage: python wu_master_template.py <analysis_data.json>")
        return 1
    with open(sys.argv[1], 'r', encoding='utf-8') as f:
        wu_data = json.load(f)
    report = generate_report(wu_data)
    print("Report generated: " + report)
    return 0


if __name__ == "__main__":
    sys.exit(main())
