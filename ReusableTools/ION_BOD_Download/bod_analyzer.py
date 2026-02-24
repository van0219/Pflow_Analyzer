#!/usr/bin/env python3
"""
BOD Data Analyzer
Loads NDJSON BOD files into SQLite for SQL-based analysis and comparison.

This tool:
1. Reads all downloaded BOD files (NDJSON format)
2. Loads them into a SQLite database
3. Optionally applies a custom SQL aggregation query (to match IPA output)
4. Compares BOD data against IPA output file(s) to verify data integrity
5. Runs analysis queries to identify data quality issues
6. Generates comparison reports

Usage (standalone - raw data analysis):
    python bod_analyzer.py --input downloads/my_investigation --output reports/analysis.xlsx

Usage (with SQL aggregation - matches IPA output):
    python bod_analyzer.py --input downloads/my_investigation --sql aggregation_query.sql --output reports/analysis.xlsx

Usage (with comparison file - RECOMMENDED):
    python bod_analyzer.py --input downloads/my_investigation --sql aggregation_query.sql --compare GLSummary.csv --output reports/comparison.xlsx

Usage (via Kiro Agent Hook):
    The agent hook will call this after downloading BODs.
"""

import os
import sys
import json
import sqlite3
import argparse
from datetime import datetime

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False
    print("Warning: pandas/openpyxl not installed. Excel report generation disabled.")

# Default paths
DEFAULT_INPUT_DIR = os.path.join(os.path.dirname(__file__), "downloads")
DEFAULT_REPORTS_DIR = os.path.join(os.path.dirname(__file__), "reports")


def create_database(db_path, table_schema=None):
    """Create SQLite database with table matching BOD structure."""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    cursor.execute("DROP TABLE IF EXISTS BODData")
    cursor.execute("DROP TABLE IF EXISTS BODDataAggregated")
    
    # Default schema for GLTotals - can be customized
    if table_schema is None:
        table_schema = """
            CREATE TABLE BODData (
                FEG TEXT,
                Scenario TEXT,
                Company TEXT,
                Ledger TEXT,
                System TEXT,
                Currency TEXT,
                EntityYearPeriod TEXT,
                Account TEXT,
                SubAccount TEXT,
                Project TEXT,
                PeriodEndingDate TEXT,
                Entity TEXT,
                FinanceDimension2 TEXT,
                Branch TEXT,
                CostCenter TEXT,
                FCLoanType TEXT,
                FinanceDimension6 TEXT,
                FinanceDimension7 TEXT,
                FinanceDimension8 TEXT,
                InterCompany TEXT,
                FinanceDimension10 TEXT,
                FunctionalAmount REAL,
                ProjectAmount REAL,
                UnitsAmount REAL,
                ADB_MTD REAL,
                ADB_QTD REAL,
                ADB_YTD REAL,
                SourceFile TEXT
            )
        """
    
    cursor.execute(table_schema)
    
    # Create indexes for common queries
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_period ON BODData(EntityYearPeriod)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_ledger ON BODData(Ledger)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_ledger_period ON BODData(Ledger, EntityYearPeriod)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_functional_amount ON BODData(FunctionalAmount)")
    
    conn.commit()
    return conn


def load_bod_files(conn, input_dir):
    """Load all NDJSON BOD files into the database."""
    cursor = conn.cursor()
    
    bod_files = [f for f in os.listdir(input_dir) if f.startswith("bod_") and f.endswith(".json")]
    bod_files.sort()
    
    if not bod_files:
        print(f"No BOD files found in {input_dir}")
        return 0
    
    print(f"Found {len(bod_files)} BOD files to load")
    
    total_records = 0
    
    for i, filename in enumerate(bod_files, 1):
        filepath = os.path.join(input_dir, filename)
        file_records = 0
        
        with open(filepath, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                
                try:
                    record = json.loads(line)
                    
                    cursor.execute("""
                        INSERT INTO BODData VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        record.get('FEG', ''),
                        record.get('Scenario', ''),
                        record.get('Company', ''),
                        record.get('Ledger', ''),
                        record.get('System', ''),
                        record.get('Currency', ''),
                        record.get('EntityYearPeriod', ''),
                        record.get('Account', ''),
                        record.get('SubAccount', ''),
                        record.get('Project', ''),
                        record.get('PeriodEndingDate', ''),
                        record.get('Entity', ''),
                        record.get('FinanceDimension2', ''),
                        record.get('Branch', ''),
                        record.get('CostCenter', ''),
                        record.get('FCLoanType', ''),
                        record.get('FinanceDimension6', ''),
                        record.get('FinanceDimension7', ''),
                        record.get('FinanceDimension8', ''),
                        record.get('InterCompany', ''),
                        record.get('FinanceDimension10', ''),
                        record.get('FunctionalAmount', 0),
                        record.get('ProjectAmount', 0),
                        record.get('UnitsAmount', 0),
                        record.get('ADB_MTD', 0),
                        record.get('ADB_QTD', 0),
                        record.get('ADB_YTD', 0),
                        filename
                    ))
                    file_records += 1
                except json.JSONDecodeError as e:
                    print(f"  Error parsing line in {filename}: {e}")
        
        total_records += file_records
        
        if i % 50 == 0 or i == len(bod_files):
            print(f"  Loaded {i}/{len(bod_files)} files ({total_records:,} records)")
            conn.commit()
    
    conn.commit()
    print(f"\nTotal records loaded: {total_records:,}")
    return total_records


def apply_sql_aggregation(conn, sql_file):
    """
    Apply a custom SQL transformation query to create BODDataAggregated table.
    
    The SQL file should contain a SELECT query that transforms the raw data.
    The query result will be stored in BODDataAggregated table.
    
    IMPORTANT: The SQL must reference 'BODData' as the source table name.
    If your original IPA SQL uses a different table name (e.g., FPI_FCE_IDL_GLTotals),
    you need to replace it with 'BODData' in your SQL file.
    
    For nested queries, only replace the innermost FROM clause that references
    the actual data source table.
    
    The SQL can include any transformations:
    - Column selection (SELECT specific columns)
    - Aggregations (GROUP BY with SUM, COUNT, etc.)
    - Field transformations (CASE statements, concatenation)
    - Null handling (IFNULL, COALESCE)
    - Column renaming (AS aliases)
    - Nested subqueries
    """
    print(f"\nApplying SQL transformation from: {sql_file}")
    
    with open(sql_file, 'r', encoding='utf-8') as f:
        sql_query = f.read().strip()
    
    # Remove any trailing semicolon
    if sql_query.endswith(';'):
        sql_query = sql_query[:-1]
    
    # Check if BODData is referenced in the query
    if 'BODData' not in sql_query and 'boddata' not in sql_query.lower():
        print("  WARNING: 'BODData' not found in SQL query.")
        print("           Make sure your SQL references 'BODData' as the source table.")
        print("           The SQLite database table is named 'BODData'.")
    
    cursor = conn.cursor()
    
    # Create aggregated table from the query
    create_sql = f"CREATE TABLE BODDataAggregated AS {sql_query}"
    
    try:
        cursor.execute(create_sql)
        conn.commit()
        
        # Count aggregated records
        cursor.execute("SELECT COUNT(*) FROM BODDataAggregated")
        agg_count = cursor.fetchone()[0]
        
        print(f"  Aggregated table created with {agg_count:,} records")
        
        # Create indexes on aggregated table
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_agg_period ON BODDataAggregated(EntityYearPeriod)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_agg_ledger ON BODDataAggregated(Ledger)")
        conn.commit()
        
        return agg_count
        
    except sqlite3.Error as e:
        print(f"  ERROR applying SQL aggregation: {e}")
        return 0


def run_analysis(conn, use_aggregated=False):
    """Run analysis queries and return results."""
    cursor = conn.cursor()
    results = {}
    
    # Determine which table to analyze
    table_name = "BODDataAggregated" if use_aggregated else "BODData"
    results['table_analyzed'] = table_name
    results['is_aggregated'] = use_aggregated
    
    # Check if table exists
    cursor.execute(f"SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name='{table_name}'")
    if cursor.fetchone()[0] == 0:
        print(f"  WARNING: Table {table_name} does not exist, falling back to BODData")
        table_name = "BODData"
        results['table_analyzed'] = table_name
        results['is_aggregated'] = False
    
    # Total record count
    cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
    results['total_records'] = cursor.fetchone()[0]
    
    # Records by Period
    cursor.execute(f"""
        SELECT EntityYearPeriod, COUNT(*) as RecordCount
        FROM {table_name}
        GROUP BY EntityYearPeriod
        ORDER BY EntityYearPeriod
    """)
    results['by_period'] = cursor.fetchall()
    
    # Records by Ledger
    cursor.execute(f"""
        SELECT Ledger, COUNT(*) as RecordCount
        FROM {table_name}
        GROUP BY Ledger
        ORDER BY RecordCount DESC
    """)
    results['by_ledger'] = cursor.fetchall()
    
    # Zero FunctionalAmount analysis by Period and Ledger
    cursor.execute(f"""
        SELECT 
            Ledger,
            EntityYearPeriod,
            COUNT(*) as TotalRecords,
            SUM(CASE WHEN FunctionalAmount = 0 THEN 1 ELSE 0 END) as ZeroRecords,
            SUM(CASE WHEN FunctionalAmount != 0 THEN 1 ELSE 0 END) as NonZeroRecords,
            ROUND(SUM(CASE WHEN FunctionalAmount = 0 THEN 1.0 ELSE 0 END) / COUNT(*) * 100, 2) as ZeroPercent
        FROM {table_name}
        GROUP BY Ledger, EntityYearPeriod
        ORDER BY Ledger, EntityYearPeriod
    """)
    results['zero_analysis'] = cursor.fetchall()
    
    # If we have both raw and aggregated, show comparison
    if use_aggregated:
        cursor.execute("SELECT COUNT(*) FROM BODData")
        results['raw_record_count'] = cursor.fetchone()[0]
        results['aggregation_ratio'] = results['raw_record_count'] / results['total_records'] if results['total_records'] > 0 else 0
    
    return results


def load_comparison_file(compare_file, conn):
    """
    Load a comparison CSV file into the database for comparison.
    
    Returns dict with comparison file stats.
    """
    if not HAS_EXCEL:
        print("  WARNING: pandas not installed, cannot load comparison file")
        return None
    
    print(f"\nLoading comparison file: {compare_file}")
    
    try:
        # Read CSV file
        df = pd.read_csv(compare_file)
        print(f"  Loaded {len(df):,} records from comparison file")
        
        # Store in database
        cursor = conn.cursor()
        cursor.execute("DROP TABLE IF EXISTS ComparisonData")
        
        # Create table and load data
        df.to_sql('ComparisonData', conn, if_exists='replace', index=False)
        
        # Create indexes
        if 'EntityYearPeriod' in df.columns:
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_comp_period ON ComparisonData(EntityYearPeriod)")
        if 'Ledger' in df.columns:
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_comp_ledger ON ComparisonData(Ledger)")
        
        conn.commit()
        
        return {
            'file': compare_file,
            'record_count': len(df),
            'columns': list(df.columns)
        }
        
    except Exception as e:
        print(f"  ERROR loading comparison file: {e}")
        return None


def run_comparison(conn, use_aggregated=False):
    """
    Compare BOD data against the comparison file.
    
    Returns comparison results dict.
    """
    cursor = conn.cursor()
    results = {}
    
    # Check if ComparisonData table exists
    cursor.execute("SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name='ComparisonData'")
    if cursor.fetchone()[0] == 0:
        print("  WARNING: No comparison data loaded")
        return None
    
    # Determine BOD table to use
    bod_table = "BODDataAggregated" if use_aggregated else "BODData"
    
    # Check if BOD table exists
    cursor.execute(f"SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name='{bod_table}'")
    if cursor.fetchone()[0] == 0:
        bod_table = "BODData"
    
    results['bod_table'] = bod_table
    
    # Get record counts
    cursor.execute(f"SELECT COUNT(*) FROM {bod_table}")
    results['bod_count'] = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM ComparisonData")
    results['comparison_count'] = cursor.fetchone()[0]
    
    # Zero analysis for comparison file (if FunctionalAmount column exists)
    cursor.execute("SELECT name FROM pragma_table_info('ComparisonData') WHERE name='FunctionalAmount'")
    has_functional_amount = cursor.fetchone() is not None
    
    if has_functional_amount:
        # Check if Ledger and EntityYearPeriod columns exist
        cursor.execute("SELECT name FROM pragma_table_info('ComparisonData') WHERE name='Ledger'")
        has_ledger = cursor.fetchone() is not None
        
        cursor.execute("SELECT name FROM pragma_table_info('ComparisonData') WHERE name='EntityYearPeriod'")
        has_period = cursor.fetchone() is not None
        
        if has_ledger and has_period:
            cursor.execute("""
                SELECT 
                    Ledger,
                    EntityYearPeriod,
                    COUNT(*) as TotalRecords,
                    SUM(CASE WHEN FunctionalAmount = 0 THEN 1 ELSE 0 END) as ZeroRecords,
                    SUM(CASE WHEN FunctionalAmount != 0 THEN 1 ELSE 0 END) as NonZeroRecords,
                    ROUND(SUM(CASE WHEN FunctionalAmount = 0 THEN 1.0 ELSE 0 END) / COUNT(*) * 100, 2) as ZeroPercent
                FROM ComparisonData
                GROUP BY Ledger, EntityYearPeriod
                ORDER BY Ledger, EntityYearPeriod
            """)
            results['comparison_zero_analysis'] = cursor.fetchall()
    
    return results


def print_comparison(results, comparison_results):
    """Print comparison results to console."""
    if not comparison_results:
        return
    
    print("\n" + "=" * 70)
    print("BOD vs IPA OUTPUT COMPARISON")
    print("=" * 70)
    
    print(f"\nBOD Data ({comparison_results['bod_table']}): {comparison_results['bod_count']:,} records")
    print(f"IPA Output (ComparisonData): {comparison_results['comparison_count']:,} records")
    
    diff = comparison_results['bod_count'] - comparison_results['comparison_count']
    if diff == 0:
        print(f"Record Count Match: YES ✓")
    else:
        print(f"Record Count Difference: {diff:+,} records")
    
    if 'comparison_zero_analysis' in comparison_results:
        print("\n" + "-" * 70)
        print("Zero FunctionalAmount Comparison (IPA Output)")
        print("-" * 70)
        print(f"  {'Ledger':<20} {'Period':<12} {'Total':>10} {'Zeros':>10} {'Non-Zero':>10} {'Zero %':>10}")
        print(f"  {'-'*20} {'-'*12} {'-'*10} {'-'*10} {'-'*10} {'-'*10}")
        for row in comparison_results['comparison_zero_analysis']:
            print(f"  {row[0]:<20} {row[1]:<12} {row[2]:>10,} {row[3]:>10,} {row[4]:>10,} {row[5]:>9.2f}%")


def print_analysis(results):
    """Print analysis results to console."""
    print("\n" + "=" * 70)
    print("BOD DATA ANALYSIS")
    print("=" * 70)
    
    if results.get('is_aggregated'):
        print(f"\n*** ANALYZING AGGREGATED DATA (matches IPA output) ***")
        print(f"Raw BOD Records: {results.get('raw_record_count', 'N/A'):,}")
        print(f"Aggregated Records: {results['total_records']:,}")
        print(f"Aggregation Ratio: {results.get('aggregation_ratio', 0):.1f}x")
    else:
        print(f"\n*** ANALYZING RAW BOD DATA ***")
        print(f"Total Records: {results['total_records']:,}")
    
    print("\n" + "-" * 70)
    print("Records by Period")
    print("-" * 70)
    for row in results['by_period']:
        print(f"  {row[0]}: {row[1]:,} records")
    
    print("\n" + "-" * 70)
    print("Records by Ledger")
    print("-" * 70)
    for row in results['by_ledger']:
        print(f"  {row[0]}: {row[1]:,} records")
    
    print("\n" + "-" * 70)
    print("Zero FunctionalAmount Analysis")
    print("-" * 70)
    print(f"  {'Ledger':<20} {'Period':<12} {'Total':>10} {'Zeros':>10} {'Non-Zero':>10} {'Zero %':>10}")
    print(f"  {'-'*20} {'-'*12} {'-'*10} {'-'*10} {'-'*10} {'-'*10}")
    for row in results['zero_analysis']:
        print(f"  {row[0]:<20} {row[1]:<12} {row[2]:>10,} {row[3]:>10,} {row[4]:>10,} {row[5]:>9.2f}%")


def generate_excel_report(results, output_path, input_dir, sql_file=None, comparison_results=None, compare_file=None):
    """Generate Excel report with analysis and comparison results."""
    if not HAS_EXCEL:
        print("Excel generation skipped - pandas/openpyxl not installed")
        return False
    
    wb = Workbook()
    
    # Styles
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    
    summary_data = [
        ["BOD Data Analysis Report"],
        [""],
        ["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Source:", input_dir],
        ["Analysis Mode:", "Aggregated (matches IPA)" if results.get('is_aggregated') else "Raw BOD Data"],
    ]
    
    if sql_file:
        summary_data.append(["SQL Aggregation:", sql_file])
    
    if compare_file:
        summary_data.append(["Comparison File:", compare_file])
    
    if results.get('is_aggregated'):
        summary_data.append(["Raw Records:", f"{results.get('raw_record_count', 'N/A'):,}"])
        summary_data.append(["Aggregated Records:", f"{results['total_records']:,}"])
        summary_data.append(["Aggregation Ratio:", f"{results.get('aggregation_ratio', 0):.1f}x"])
    else:
        summary_data.append(["Total Records:", f"{results['total_records']:,}"])
    
    # Add comparison summary if available
    if comparison_results:
        summary_data.append([""])
        summary_data.append(["=== COMPARISON RESULTS ==="])
        summary_data.append(["BOD Records:", f"{comparison_results['bod_count']:,}"])
        summary_data.append(["IPA Output Records:", f"{comparison_results['comparison_count']:,}"])
        diff = comparison_results['bod_count'] - comparison_results['comparison_count']
        if diff == 0:
            summary_data.append(["Record Count Match:", "YES ✓"])
        else:
            summary_data.append(["Record Count Difference:", f"{diff:+,}"])
    
    summary_data.append([""])
    summary_data.append(["Records by Period:"])
    
    for row in results['by_period']:
        summary_data.append([f"  {row[0]}", f"{row[1]:,}"])
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            if row_idx == 5 and results.get('is_aggregated'):
                cell.fill = green_fill
            if "YES ✓" in str(value):
                cell.fill = green_fill
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    
    # Zero Analysis sheet (BOD Data)
    ws2 = wb.create_sheet("BOD Zero Analysis")
    
    headers = ["Ledger", "Period", "Total", "Zeros", "Non-Zero", "Zero %"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    for row_idx, row_data in enumerate(results['zero_analysis'], 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            # Highlight high zero percentages
            if col_idx == 6 and isinstance(value, (int, float)) and value > 50:
                cell.fill = highlight_fill
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws2.column_dimensions[col].width = 15
    
    # Comparison Zero Analysis sheet (if comparison data available)
    if comparison_results and 'comparison_zero_analysis' in comparison_results:
        ws3 = wb.create_sheet("IPA Output Zero Analysis")
        
        headers = ["Ledger", "Period", "Total", "Zeros", "Non-Zero", "Zero %"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws3.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        
        for row_idx, row_data in enumerate(comparison_results['comparison_zero_analysis'], 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                # Highlight high zero percentages
                if col_idx == 6 and isinstance(value, (int, float)) and value > 50:
                    cell.fill = highlight_fill
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws3.column_dimensions[col].width = 15
    
    wb.save(output_path)
    print(f"\nExcel report saved to: {output_path}")
    return True


def analyze_bods(input_dir, output_report=None, sql_file=None, compare_file=None):
    """
    Main function to analyze BOD data and compare against IPA output.
    
    Args:
        input_dir: Directory containing downloaded BOD files
        output_report: Path for Excel report output
        sql_file: Optional SQL file with transformation query to match IPA output.
                  If not provided, will check for sql/transformation.sql in input_dir.
        compare_file: REQUIRED for complete analysis. Path to IPA output file (e.g., GLSummary.csv)
                      to compare against BOD data. If not provided, will check for CSV files
                      in the comparison/ subfolder of input_dir.
    """
    print("=" * 70)
    print("BOD Data Analyzer")
    print("=" * 70)
    print(f"Input: {input_dir}")
    
    # Check for SQL file in investigation's sql/ subfolder if not explicitly provided
    if not sql_file:
        default_sql_path = os.path.join(input_dir, "sql", "transformation.sql")
        if os.path.exists(default_sql_path):
            sql_file = default_sql_path
            print(f"Found SQL transformation file: {sql_file}")
    
    if sql_file:
        print(f"SQL Transformation: {sql_file}")
    
    # Check for comparison file in investigation folder if not explicitly provided
    if not compare_file:
        # First check for CSV files directly in the investigation folder
        csv_files_in_root = [f for f in os.listdir(input_dir) if f.endswith('.csv') and not f.startswith('bod_')]
        if csv_files_in_root:
            compare_file = os.path.join(input_dir, csv_files_in_root[0])
            print(f"Found comparison file: {compare_file}")
            if len(csv_files_in_root) > 1:
                print(f"  NOTE: Multiple CSV files found, using first one: {csv_files_in_root[0]}")
                print(f"  Other files: {', '.join(csv_files_in_root[1:])}")
        else:
            # Also check comparison/ subfolder for backward compatibility
            comparison_dir = os.path.join(input_dir, "comparison")
            if os.path.exists(comparison_dir):
                csv_files = [f for f in os.listdir(comparison_dir) if f.endswith('.csv')]
                if csv_files:
                    compare_file = os.path.join(comparison_dir, csv_files[0])
                    print(f"Found comparison file: {compare_file}")
                    if len(csv_files) > 1:
                        print(f"  NOTE: Multiple CSV files found, using first one: {csv_files[0]}")
                        print(f"  Other files: {', '.join(csv_files[1:])}")
    
    if compare_file:
        print(f"Comparison File: {compare_file}")
    else:
        print("\n*** WARNING: No comparison file provided ***")
        print("  Analysis will only show BOD data without IPA output comparison.")
        print("  For complete analysis, provide a comparison file using --compare")
        print("  or place CSV file(s) in the comparison/ subfolder.\n")
    
    # Create database in input directory
    db_path = os.path.join(input_dir, "bod_data.db")
    print(f"Database: {db_path}")
    
    # Create and load database
    print("\nCreating database...")
    conn = create_database(db_path)
    
    print("\nLoading BOD files...")
    total = load_bod_files(conn, input_dir)
    
    if total == 0:
        conn.close()
        return {'success': False, 'error': 'No records loaded'}
    
    # Apply SQL aggregation if provided
    use_aggregated = False
    if sql_file and os.path.exists(sql_file):
        agg_count = apply_sql_aggregation(conn, sql_file)
        if agg_count > 0:
            use_aggregated = True
    
    # Run analysis on BOD data
    print("\nRunning BOD analysis...")
    results = run_analysis(conn, use_aggregated=use_aggregated)
    
    # Print BOD analysis to console
    print_analysis(results)
    
    # Load and compare against IPA output if provided
    comparison_results = None
    comparison_info = None
    if compare_file and os.path.exists(compare_file):
        comparison_info = load_comparison_file(compare_file, conn)
        if comparison_info:
            comparison_results = run_comparison(conn, use_aggregated=use_aggregated)
            print_comparison(results, comparison_results)
    
    # Generate Excel report if requested
    if output_report:
        os.makedirs(os.path.dirname(output_report), exist_ok=True)
        generate_excel_report(
            results, 
            output_report, 
            input_dir, 
            sql_file, 
            comparison_results=comparison_results,
            compare_file=compare_file
        )
    
    conn.close()
    
    print("\n" + "=" * 70)
    print("ANALYSIS COMPLETE")
    print("=" * 70)
    print(f"Database: {db_path}")
    print("You can run additional SQL queries using: sqlite3 " + db_path)
    if use_aggregated:
        print("  - BODData: Raw BOD records")
        print("  - BODDataAggregated: Aggregated records (matches IPA output)")
    if comparison_results:
        print("  - ComparisonData: IPA output records for comparison")
    
    return {
        'success': True,
        'total_records': results['total_records'],
        'is_aggregated': use_aggregated,
        'db_path': db_path,
        'report_path': output_report,
        'comparison_file': compare_file,
        'comparison_results': comparison_results
    }


def main():
    """CLI entry point."""
    parser = argparse.ArgumentParser(
        description='Analyze downloaded BOD data and compare against IPA output',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Basic analysis (BOD data only)
  python bod_analyzer.py --input GLSummary_Investigation

  # With SQL aggregation (to match IPA output structure)
  python bod_analyzer.py --input GLSummary_Investigation --sql sql/transformation.sql

  # With comparison file (RECOMMENDED - complete analysis)
  python bod_analyzer.py --input GLSummary_Investigation --compare comparison/GLSummary.csv

  # Full analysis with all options
  python bod_analyzer.py --input GLSummary_Investigation --sql sql/transformation.sql --compare comparison/GLSummary.csv --output reports/analysis.xlsx

Note: If --sql or --compare are not provided, the tool will look for:
  - sql/transformation.sql in the input directory
  - CSV files in the comparison/ subfolder of the input directory
        """
    )
    parser.add_argument('--input', default=DEFAULT_INPUT_DIR, 
                        help='Input directory with BOD files (investigation folder)')
    parser.add_argument('--output', 
                        help='Output Excel report path (optional, auto-generated if not specified)')
    parser.add_argument('--sql', 
                        help='SQL file with transformation query to match IPA output (optional). The query should SELECT from BODData table.')
    parser.add_argument('--compare', 
                        help='IPA output CSV file to compare against BOD data (RECOMMENDED). This proves whether zeros came from source or were introduced by IPA.')
    
    args = parser.parse_args()
    
    output_report = args.output
    if not output_report:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_report = os.path.join(DEFAULT_REPORTS_DIR, f"bod_analysis_{timestamp}.xlsx")
    
    result = analyze_bods(args.input, output_report, args.sql, args.compare)
    
    return 0 if result.get('success') else 1


if __name__ == "__main__":
    sys.exit(main())
